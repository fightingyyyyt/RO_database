#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
膜材料数据库浓度清洗脚本（LLM + 规则 + 可回溯骨架版）

用途
----
针对以下 5 类字段做逐槽位（slot）判定与标准化：
1. 水相单体 / 水相单体浓度 / 水相单体浓度_单位
2. 油相单体 / 油相单体浓度 / 油相单体浓度_单位
3. 添加剂 / 添加剂浓度 / 添加剂浓度_单位
4. 改性剂 / 改性剂浓度 / 改性剂浓度_单位
5. 测试NaCl浓度 / 测试NaCl浓度_单位

核心思想
--------
- 一行数据拆成 5 个 slot，而不是整行一次性判断。
- 先做结构化解析，再调用 LLM 做语义判断。
- 真正的数值换算尽量交给程序；LLM 负责判断“能不能转 / 还缺什么条件 / 这是什么相 / % 更像什么”。
- 不能可靠换算时，打标，而不是硬转。

说明
----
- 默认使用 DeepSeek API（接口兼容 OpenAI 规范，可切换其他服务）
- 支持 --mock-llm 模式：不用联网，也能把整条流程先跑通
- 输出：
  1) 主结果表 cleaned_main
  2) 异常队列表 review_queue
  3) 长表 slot_results_long
  4) summary 统计表

环境变量（默认使用 DeepSeek）
--------
DEEPSEEK_API_KEY=...                          # 填 DeepSeek API Key
DEEPSEEK_BASE_URL=https://api.deepseek.com/v1  # 默认 DeepSeek；可改为其他兼容地址
DEEPSEEK_MODEL=deepseek-chat                   # 或 deepseek-reasoner 等

运行方式
--------
在脚本内「0. 文档与运行配置」中配置好 input_file、output_file、sheet_name、limit、mock_llm 等
（文件路径为与 py 同目录下的文件名即可），然后直接运行，无需在终端输入路径：

  python membrane_concentration_cleaner.py

若需临时覆盖，仍可使用命令行参数，例如：
  python membrane_concentration_cleaner.py --limit 100 --mock-llm
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from openai import OpenAI
except Exception:
    OpenAI = None  # type: ignore

# =========================
# 0. 文档与运行配置（与 py 同目录，无需在终端输入）
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent  # 本脚本所在目录

RUN_CONFIG = {
    "input_file": "test.xlsx",           # 输入 Excel 文件名（与 py 同目录）
    "output_file": "test_output.xlsx",  # 输出 Excel 文件名（与 py 同目录）
    "sheet_name": None,                  # 工作表名，None 表示用第一个 sheet
    "limit": None,                      # 只处理前 N 行，None 表示全部
    "mock_llm": False,                  # True=不调用 DeepSeek 仅 mock，False=真实调用
}

# =========================
# 1. 常量与状态码
# =========================

SLOT_CONFIG = [
    {
        "slot_name": "aqueous_monomer",
        "label_cn": "水相单体",
        "solute_col": "水相单体",
        "value_col": "水相单体浓度",
        "unit_col": "水相单体浓度_单位",
        "phase_hint": "aqueous",
        "fixed_solute": None,
    },
    {
        "slot_name": "organic_monomer",
        "label_cn": "油相单体",
        "solute_col": "油相单体",
        "value_col": "油相单体浓度",
        "unit_col": "油相单体浓度_单位",
        "phase_hint": "organic",
        "fixed_solute": None,
    },
    {
        "slot_name": "additive",
        "label_cn": "添加剂",
        "solute_col": "添加剂",
        "value_col": "添加剂浓度",
        "unit_col": "添加剂浓度_单位",
        "phase_hint": "unknown",
        "fixed_solute": None,
    },
    {
        "slot_name": "modifier",
        "label_cn": "改性剂",
        "solute_col": "改性剂",
        "value_col": "改性剂浓度",
        "unit_col": "改性剂浓度_单位",
        "phase_hint": "unknown",
        "fixed_solute": None,
    },
    {
        "slot_name": "test_nacl",
        "label_cn": "测试NaCl",
        "solute_col": None,
        "value_col": "测试NaCl浓度",
        "unit_col": "测试NaCl浓度_单位",
        "phase_hint": "test_aqueous",
        "fixed_solute": "NaCl",
    },
]

PARSE_STATUS = {
    "OK",
    "EMPTY_ALL",
    "EMPTY_VALUE",
    "EMPTY_UNIT",
    "EMPTY_SOLUTE",
    "NON_NUMERIC_VALUE",
    "MULTI_COUNT_MISMATCH",
    "UNSUPPORTED_FORMAT",
}

MAPPING_STATUS = {
    "SINGLE_ITEM",
    "ONE_TO_ONE",
    "ONE_VALUE_MULTI_SOLUTE",
    "ONE_UNIT_MULTI_SOLUTE",
    "MULTI_UNRESOLVED",
}

PHASE_STATUS = {"aqueous", "organic", "test_aqueous", "mixed", "unknown"}
PHASE_EVIDENCE = {"column_hint", "row_context", "source_text", "chemical_knowledge", "unknown"}
PERCENT_TYPE = {"WT_PERCENT", "VOL_PERCENT", "WV_PERCENT", "MOL_PERCENT", "PERCENT_UNKNOWN", "NOT_PERCENT_UNIT"}
WTPERCENT_STATUS = {"DIRECT_WT", "SAFE_CONVERTED", "ASSUMED_CONVERTED", "CANNOT_CONVERT", "NEED_TRACEBACK", "FAILED_PARSE"}
CONFIDENCE = {"HIGH", "MEDIUM", "LOW"}
TRACEBACK_TARGET_ALLOWED = {
    "SOLUTE_IDENTITY",
    "PHASE_TYPE",
    "PERCENT_TYPE",
    "SOLVENT_TYPE",
    "MOLECULAR_WEIGHT",
    "DENSITY",
    "MAPPING_RELATION",
    "ORIGINAL_UNIT",
}

DIRECT_WT_UNITS = {
    "wt%", "wt.%", "w/w%", "mass%", "mass %", "g/100g", "g per 100 g", "重量%", "质量分数%"
}

PERCENT_ONLY_UNITS = {"%", "％", "percent", "percentage"}
MOLAR_UNITS = {"m", "mol/l", "mol l-1", "mol·l-1", "mol/L", "M", "mM", "mmol/L"}
PPM_UNITS = {"ppm", "ppm nacl"}
GL_UNITS = {"g/l", "g/L"}
MGL_UNITS = {"mg/l", "mg/L"}
VOL_UNITS = {"vol%", "v/v%", "volume%", "volume %"}
WV_UNITS = {"w/v%", "g/100ml", "g/100mL"}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
OK_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

# =========================
# 2. Prompt 模板
# =========================

SYSTEM_PROMPT = r"""
You are an expert data-cleaning assistant for membrane-material literature databases.

Your job is to normalize one concentration slot from one experimental record.

You must determine, using only the provided row context:
1. what the solute is,
2. what phase it belongs to,
3. whether the numeric value is valid,
4. what the unit means,
5. whether the concentration can be converted to wt%,
6. whether traceback to original source text is needed.

Important rules:
- Do not guess missing facts.
- Do not invent molecular weight, density, solvent, or concentration meaning unless clearly supported by the provided context.
- If the evidence is insufficient, mark need_traceback=true or wtpercent_status=CANNOT_CONVERT / NEED_TRACEBACK.
- If the original unit is only "%" and its meaning is ambiguous, do not assume wt% unless the context strongly supports it.
- For test NaCl concentration, the solute is NaCl unless explicitly contradicted.
- For aqueous monomer slot, prefer the aqueous monomer column as the solute.
- For organic monomer slot, prefer the organic monomer column as the solute.
- For additive and modifier slots, use row context to infer solute-phase relation, but do not over-infer.
- Keep parsed_value numeric only if the raw value can be reliably parsed as a number.
- If multiple solutes / values / units are present, determine whether mapping is one-to-one. If mapping is unclear, mark MULTI_UNRESOLVED.
- Output JSON only.
- Do not output markdown.
- All enum values must exactly match the allowed values.

Strict anti-hallucination rule:
If a required fact is not explicitly stated or cannot be strongly inferred from the provided row context, do not fabricate it.
Prefer null, NEED_TRACEBACK, CANNOT_CONVERT, or PERCENT_UNKNOWN over unsupported completion.

Allowed enum values:

parse_status:
OK
EMPTY_ALL
EMPTY_VALUE
EMPTY_UNIT
EMPTY_SOLUTE
NON_NUMERIC_VALUE
MULTI_COUNT_MISMATCH
UNSUPPORTED_FORMAT

mapping_status:
SINGLE_ITEM
ONE_TO_ONE
ONE_VALUE_MULTI_SOLUTE
ONE_UNIT_MULTI_SOLUTE
MULTI_UNRESOLVED

phase_identified:
aqueous
organic
test_aqueous
mixed
unknown

phase_evidence:
column_hint
row_context
source_text
chemical_knowledge
unknown

percent_type_inferred:
WT_PERCENT
VOL_PERCENT
WV_PERCENT
MOL_PERCENT
PERCENT_UNKNOWN
NOT_PERCENT_UNIT

wtpercent_status:
DIRECT_WT
SAFE_CONVERTED
ASSUMED_CONVERTED
CANNOT_CONVERT
NEED_TRACEBACK
FAILED_PARSE

confidence:
HIGH
MEDIUM
LOW

review_flag:
0
1
2

traceback_target allowed items:
SOLUTE_IDENTITY
PHASE_TYPE
PERCENT_TYPE
SOLVENT_TYPE
MOLECULAR_WEIGHT
DENSITY
MAPPING_RELATION
ORIGINAL_UNIT
""".strip()

USER_PROMPT_TEMPLATE = r"""
Normalize the following concentration slot from one membrane experiment record.

Return one JSON object only.

Row context:
{row_context_json}

Current slot:
{current_slot_json}

Required output schema:
{{
  "slot_name": "string",
  "slot_item_index": 1,
  "parse_status": "enum",
  "mapping_status": "enum",
  "raw_solute": "string or null",
  "raw_value": "string or null",
  "raw_unit": "string or null",
  "parsed_solute": "string or null",
  "parsed_solute_alias": "string or null",
  "parsed_value": "number or null",
  "parsed_unit": "string or null",
  "phase_identified": "enum",
  "phase_evidence": "enum",
  "solvent_identified": "string or null",
  "percent_type_inferred": "enum",
  "can_convert_to_wtpercent": true,
  "wtpercent_value": "number or null",
  "wtpercent_status": "enum",
  "conversion_method": "string or null",
  "requires_molecular_weight": true,
  "requires_density": true,
  "molecular_weight_value": "number or null",
  "molecular_weight_source": "string or null",
  "density_value": "number or null",
  "density_source": "string or null",
  "need_traceback": true,
  "traceback_target": [],
  "review_flag": 0,
  "confidence": "enum",
  "reason": "string"
}}

Additional decision rules:
- If parse_status is not OK, then wtpercent_status should usually be FAILED_PARSE.
- If the unit is already wt%, w/w%, mass%, or equivalent mass fraction, use DIRECT_WT.
- If conversion is impossible from current evidence, use CANNOT_CONVERT or NEED_TRACEBACK.
- Do not fill molecular_weight_value or density_value unless explicitly present in the provided context.
- If the unit is only "%" and context is insufficient, set percent_type_inferred=PERCENT_UNKNOWN.
- review_flag=0 for DIRECT_WT or SAFE_CONVERTED with strong evidence.
- review_flag=1 for ASSUMED_CONVERTED.
- review_flag=2 for FAILED_PARSE, CANNOT_CONVERT, or NEED_TRACEBACK.
- Keep reason concise, ideally within 1 sentence.

Slot-specific hint:
{slot_specific_hint}

Return valid JSON only. No explanation outside JSON.
""".strip()

SLOT_HINTS = {
    "aqueous_monomer": (
        "For aqueous_monomer, prefer the aqueous monomer column as the solute identity. "
        "Unless strong contrary evidence exists, phase_identified should be aqueous. "
        "If raw_unit is '%' only, infer WT_PERCENT only when the context strongly indicates aqueous formulation mass fraction."
    ),
    "organic_monomer": (
        "For organic_monomer, prefer the organic monomer column as the solute identity. "
        "Unless strong contrary evidence exists, phase_identified should be organic. "
        "Use organic_solvent context if available. "
        "Do not assume '%' means wt% unless the formulation context supports mass fraction in organic phase."
    ),
    "additive": (
        "For additive, determine whether the additive belongs to aqueous phase, organic phase, mixed phase, or unknown based on row context. "
        "If multiple additives / values / units exist, determine whether they can be mapped one-to-one. "
        "If mapping is unclear, use mapping_status=MULTI_UNRESOLVED and need_traceback=true. "
        "Do not assume all additives are in the same phase unless supported by context."
    ),
    "modifier": (
        "For modifier, be conservative. "
        "A modifier may belong to post-treatment solution, coating solution, grafting solution, or another modification system. "
        "Do not force phase_identified to aqueous or organic unless supported by context. "
        "If the context is insufficient to determine the concentration meaning, set need_traceback=true."
    ),
    "test_nacl": (
        "For test_nacl, the solute is NaCl unless explicitly contradicted. "
        "Phase is usually test_aqueous. "
        "Units like ppm, mg/L, g/L, mol/L, and % should be interpreted as NaCl concentration in feed/test solution. "
        "If only dilute-aqueous approximation would allow conversion to wt%, mark ASSUMED_CONVERTED rather than SAFE_CONVERTED."
    ),
}

# =========================
# 3. 数据结构
# =========================

@dataclass
class SlotResult:
    row_index: int
    row_id: str
    slot_name: str
    slot_item_index: int
    parse_status: str
    mapping_status: str
    raw_solute: Optional[str]
    raw_value: Optional[str]
    raw_unit: Optional[str]
    parsed_solute: Optional[str]
    parsed_solute_alias: Optional[str]
    parsed_value: Optional[float]
    parsed_unit: Optional[str]
    phase_identified: str
    phase_evidence: str
    solvent_identified: Optional[str]
    percent_type_inferred: str
    can_convert_to_wtpercent: bool
    wtpercent_value: Optional[float]
    wtpercent_status: str
    conversion_method: Optional[str]
    requires_molecular_weight: bool
    requires_density: bool
    molecular_weight_value: Optional[float]
    molecular_weight_source: Optional[str]
    density_value: Optional[float]
    density_source: Optional[str]
    need_traceback: bool
    traceback_target: List[str]
    review_flag: int
    confidence: str
    reason: str

# =========================
# 4. 基础工具函数
# =========================

def is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    if str(v).strip() == "":
        return True
    return False


def norm_text(v: Any) -> Optional[str]:
    if is_blank(v):
        return None
    s = str(v).strip()
    s = s.replace("\u3000", " ")
    return s


def clean_unit(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    s = u.strip()
    s = s.replace("％", "%")
    s = s.replace("·", "/")
    s = re.sub(r"\s+", " ", s)
    return s


def canonical_unit(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    s = clean_unit(u)
    if s is None:
        return None
    s_low = s.lower().strip()
    # 直接标准化一些常见别名
    if s_low in {x.lower() for x in DIRECT_WT_UNITS}:
        return "wt%"
    if s_low in {x.lower() for x in PERCENT_ONLY_UNITS}:
        return "%"
    if s_low in {x.lower() for x in PPM_UNITS}:
        return "ppm"
    if s_low in {x.lower() for x in GL_UNITS}:
        return "g/L"
    if s_low in {x.lower() for x in MGL_UNITS}:
        return "mg/L"
    if s_low in {x.lower() for x in VOL_UNITS}:
        return "v/v%"
    if s_low in {x.lower() for x in WV_UNITS}:
        return "w/v%"
    if s_low in {"m", "mol/l", "mol l-1", "mol·l-1"}:
        return "mol/L"
    if s_low == "mm":
        return "mm"
    if s == "M":
        return "M"
    if s == "mM":
        return "mM"
    return s


def split_multi(v: Optional[str]) -> List[str]:
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = re.split(r"[;；]+", s)
    parts = [p.strip() for p in parts if p.strip()]
    return parts


def is_supported_numeric_format(s: str) -> bool:
    s = s.strip()
    # 支持单个纯数字、科学计数法、区间暂不支持
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s))


def to_float_safe(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    t = s.strip()
    if not is_supported_numeric_format(t):
        return None
    try:
        return float(t)
    except Exception:
        return None


def build_row_id(row: pd.Series, idx: int) -> str:
    for col in ["样品编号", "DOI", "文件名称", "论文题目"]:
        if col in row.index and not is_blank(row.get(col)):
            return str(row.get(col))
    return f"row_{idx}"


def row_context_dict(row: pd.Series, row_id: str) -> Dict[str, Any]:
    return {
        "row_id": row_id,
        "membrane_type": norm_text(row.get("膜类型")),
        "membrane_material": norm_text(row.get("膜材料")),
        "membrane_structure": norm_text(row.get("膜结构标签")) if "膜结构标签" in row.index else None,
        "aqueous_monomer": norm_text(row.get("水相单体")),
        "aqueous_monomer_value": norm_text(row.get("水相单体浓度")),
        "aqueous_monomer_unit": norm_text(row.get("水相单体浓度_单位")),
        "organic_monomer": norm_text(row.get("油相单体")),
        "organic_monomer_value": norm_text(row.get("油相单体浓度")),
        "organic_monomer_unit": norm_text(row.get("油相单体浓度_单位")),
        "additive": norm_text(row.get("添加剂")),
        "additive_value": norm_text(row.get("添加剂浓度")),
        "additive_unit": norm_text(row.get("添加剂浓度_单位")),
        "modifier": norm_text(row.get("改性剂")),
        "modifier_value": norm_text(row.get("改性剂浓度")),
        "modifier_unit": norm_text(row.get("改性剂浓度_单位")),
        "test_nacl_value": norm_text(row.get("测试NaCl浓度")),
        "test_nacl_unit": norm_text(row.get("测试NaCl浓度_单位")),
        "organic_solvent": norm_text(row.get("有机溶剂")),
        "aqueous_solvent": norm_text(row.get("水相溶剂")) if "水相溶剂" in row.index else None,
        "other_context": build_other_context(row),
    }


def build_other_context(row: pd.Series) -> Optional[str]:
    parts = []
    for col in [
        "制备方法",
        "过滤模式",
        "pH/碱度",
        "基底",
        "文件名称",
        "发表年份",
        "DOI",
        "论文题目",
    ]:
        if col in row.index and not is_blank(row.get(col)):
            parts.append(f"{col}: {row.get(col)}")
    return " | ".join(parts) if parts else None


def make_empty_result(row_index: int, row_id: str, slot_name: str, item_idx: int,
                      raw_solute: Optional[str], raw_value: Optional[str], raw_unit: Optional[str]) -> SlotResult:
    return SlotResult(
        row_index=row_index,
        row_id=row_id,
        slot_name=slot_name,
        slot_item_index=item_idx,
        parse_status="EMPTY_ALL",
        mapping_status="SINGLE_ITEM",
        raw_solute=raw_solute,
        raw_value=raw_value,
        raw_unit=raw_unit,
        parsed_solute=None,
        parsed_solute_alias=None,
        parsed_value=None,
        parsed_unit=None,
        phase_identified="unknown",
        phase_evidence="unknown",
        solvent_identified=None,
        percent_type_inferred="NOT_PERCENT_UNIT",
        can_convert_to_wtpercent=False,
        wtpercent_value=None,
        wtpercent_status="FAILED_PARSE",
        conversion_method=None,
        requires_molecular_weight=False,
        requires_density=False,
        molecular_weight_value=None,
        molecular_weight_source=None,
        density_value=None,
        density_source=None,
        need_traceback=False,
        traceback_target=[],
        review_flag=0,
        confidence="LOW",
        reason="empty slot",
    )

# =========================
# 5. 多值拆分与槽位展开
# =========================

def explode_slot_items(row: pd.Series, slot_cfg: Dict[str, Any], row_index: int, row_id: str) -> List[Dict[str, Any]]:
    slot_name = slot_cfg["slot_name"]
    fixed_solute = slot_cfg["fixed_solute"]

    raw_solute = fixed_solute if fixed_solute else norm_text(row.get(slot_cfg["solute_col"]))
    raw_value = norm_text(row.get(slot_cfg["value_col"]))
    raw_unit = norm_text(row.get(slot_cfg["unit_col"]))

    if all(is_blank(x) for x in [raw_solute, raw_value, raw_unit]):
        return [{
            "slot_name": slot_name,
            "slot_item_index": 1,
            "raw_solute": raw_solute,
            "raw_value": raw_value,
            "raw_unit": raw_unit,
            "solute_count": 0,
            "value_count": 0,
            "unit_count": 0,
            "mapping_status": "SINGLE_ITEM",
        }]

    solutes = split_multi(raw_solute) if raw_solute else []
    values = split_multi(raw_value) if raw_value else []
    units = split_multi(raw_unit) if raw_unit else []

    # test_nacl 之类 fixed_solute 单一，但 value/unit 可能多值；通常按单值处理
    if fixed_solute and not solutes:
        solutes = [fixed_solute]

    s_count = len(solutes)
    v_count = len(values)
    u_count = len(units)

    # 单值情形
    if max(s_count, v_count, u_count) <= 1:
        return [{
            "slot_name": slot_name,
            "slot_item_index": 1,
            "raw_solute": solutes[0] if solutes else raw_solute,
            "raw_value": values[0] if values else raw_value,
            "raw_unit": units[0] if units else raw_unit,
            "solute_count": s_count,
            "value_count": v_count,
            "unit_count": u_count,
            "mapping_status": "SINGLE_ITEM",
        }]

    # 完全一一对应
    if s_count == v_count == u_count and s_count > 1:
        items = []
        for i in range(s_count):
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[i],
                "raw_unit": units[i],
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_TO_ONE",
            })
        return items

    # 一值多物质
    if s_count > 1 and v_count == 1 and u_count in {1, s_count}:
        items = []
        for i in range(s_count):
            item_unit = units[i] if u_count == s_count else (units[0] if u_count == 1 else raw_unit)
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[0],
                "raw_unit": item_unit,
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_VALUE_MULTI_SOLUTE",
            })
        return items

    # 一单位多物质
    if s_count > 1 and u_count == 1 and v_count == s_count:
        items = []
        for i in range(s_count):
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[i],
                "raw_unit": units[0],
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_UNIT_MULTI_SOLUTE",
            })
        return items

    # 无法解析对应关系，也输出最小槽位，供 LLM/异常队列使用
    return [{
        "slot_name": slot_name,
        "slot_item_index": 1,
        "raw_solute": raw_solute,
        "raw_value": raw_value,
        "raw_unit": raw_unit,
        "solute_count": s_count,
        "value_count": v_count,
        "unit_count": u_count,
        "mapping_status": "MULTI_UNRESOLVED",
    }]

# =========================
# 6. Mock LLM（先把流程跑通）
# =========================

def infer_phase_by_slot(slot_name: str, row_ctx: Dict[str, Any]) -> Tuple[str, str, Optional[str]]:
    if slot_name == "aqueous_monomer":
        return "aqueous", "column_hint", (row_ctx.get("aqueous_solvent") or "water")
    if slot_name == "organic_monomer":
        return "organic", "column_hint", row_ctx.get("organic_solvent")
    if slot_name == "test_nacl":
        return "test_aqueous", "column_hint", "water"

    # addtive / modifier 尝试根据制备方法和上下文做保守推断
    other_context = (row_ctx.get("other_context") or "").lower()
    organic_solvent = (row_ctx.get("organic_solvent") or "")
    if "aqueous" in other_context or "water" in other_context or "水相" in other_context:
        return "aqueous", "row_context", (row_ctx.get("aqueous_solvent") or "water")
    if organic_solvent:
        return "organic", "row_context", organic_solvent
    return "unknown", "unknown", None


def mock_judge_slot(row_index: int, row_id: str, row_ctx: Dict[str, Any], slot_item: Dict[str, Any]) -> SlotResult:
    slot_name = slot_item["slot_name"]
    raw_solute = norm_text(slot_item.get("raw_solute"))
    raw_value = norm_text(slot_item.get("raw_value"))
    raw_unit = norm_text(slot_item.get("raw_unit"))
    mapping_status = slot_item.get("mapping_status", "SINGLE_ITEM")

    res = make_empty_result(row_index, row_id, slot_name, slot_item.get("slot_item_index", 1), raw_solute, raw_value, raw_unit)
    res.mapping_status = mapping_status

    if all(is_blank(x) for x in [raw_solute, raw_value, raw_unit]):
        res.parse_status = "EMPTY_ALL"
        res.reason = "empty slot"
        res.review_flag = 0
        return res

    # 空值判断
    if slot_name != "test_nacl" and is_blank(raw_solute):
        res.parse_status = "EMPTY_SOLUTE"
        res.reason = "solute missing"
        res.review_flag = 2
        return res
    if is_blank(raw_value):
        res.parse_status = "EMPTY_VALUE"
        res.reason = "value missing"
        res.review_flag = 1 if raw_solute else 0
        return res
    if is_blank(raw_unit):
        res.parse_status = "EMPTY_UNIT"
        res.reason = "unit missing"
        res.review_flag = 2
        return res
    if mapping_status == "MULTI_UNRESOLVED":
        res.parse_status = "MULTI_COUNT_MISMATCH"
        res.wtpercent_status = "FAILED_PARSE"
        res.need_traceback = True
        res.traceback_target = ["MAPPING_RELATION"]
        res.review_flag = 2
        res.reason = "multi-value mapping unresolved"
        return res

    parsed_num = to_float_safe(raw_value)
    unit_std = canonical_unit(raw_unit)
    phase_identified, phase_evidence, solvent_identified = infer_phase_by_slot(slot_name, row_ctx)

    res.raw_solute = raw_solute
    res.raw_value = raw_value
    res.raw_unit = raw_unit
    res.parsed_solute = raw_solute if raw_solute else ("NaCl" if slot_name == "test_nacl" else None)
    res.parsed_solute_alias = raw_solute
    res.parsed_value = parsed_num
    res.parsed_unit = unit_std
    res.phase_identified = phase_identified
    res.phase_evidence = phase_evidence
    res.solvent_identified = solvent_identified

    if parsed_num is None:
        # 比如 trace、1:4、few drops
        res.parse_status = "UNSUPPORTED_FORMAT" if re.search(r"[:/]|trace|few|excess|small amount|saturated", raw_value.lower()) else "NON_NUMERIC_VALUE"
        res.wtpercent_status = "FAILED_PARSE"
        res.review_flag = 2
        res.reason = "value is not reliably numeric"
        return res

    res.parse_status = "OK"

    # 1) 已是 wt%
    if unit_std == "wt%":
        res.percent_type_inferred = "WT_PERCENT"
        res.can_convert_to_wtpercent = True
        res.wtpercent_value = parsed_num
        res.wtpercent_status = "DIRECT_WT"
        res.conversion_method = "direct_standardization"
        res.review_flag = 0
        res.confidence = "HIGH"
        res.reason = "original unit already expresses mass fraction"
        return res

    # 2) 单独的 % —— 保守处理
    if unit_std == "%":
        # 对测试NaCl，通常默认质量分数不够稳，保守打标
        if slot_name == "test_nacl":
            res.percent_type_inferred = "PERCENT_UNKNOWN"
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PERCENT_TYPE", "ORIGINAL_UNIT"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "plain percent is ambiguous without source context"
            return res

        # 对水相/油相单体，如果列名明确是浓度且语境强，给一个保守中间态
        if slot_name in {"aqueous_monomer", "organic_monomer"}:
            res.percent_type_inferred = "PERCENT_UNKNOWN"
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PERCENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "plain percent may be wt% but source confirmation is needed"
            return res

        res.percent_type_inferred = "PERCENT_UNKNOWN"
        res.can_convert_to_wtpercent = False
        res.wtpercent_status = "NEED_TRACEBACK"
        res.need_traceback = True
        res.traceback_target = ["PERCENT_TYPE", "PHASE_TYPE"]
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "plain percent is ambiguous"
        return res

    # 3) ppm
    if unit_std == "ppm":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10000.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "ppm_to_wtpercent_dilute_aqueous_approx"
            res.requires_density = False
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "ppm converted using dilute aqueous approximation"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "ppm conversion needs phase/solution context"
        return res

    # 4) g/L or mg/L —— 仅在水相/测试液里用 rho≈1 的保守近似
    if unit_std == "g/L":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "g_per_L_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "g/L converted assuming density≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "DENSITY", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "g/L conversion requires density and phase context"
        return res

    if unit_std == "mg/L":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10000.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "mg_per_L_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "mg/L converted assuming density≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "DENSITY", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "mg/L conversion requires density and phase context"
        return res

    # 5) 摩尔浓度
    if unit_std in {"M", "mM", "mol/L"}:
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        res.can_convert_to_wtpercent = False
        res.requires_molecular_weight = True
        res.requires_density = True
        res.wtpercent_status = "NEED_TRACEBACK"
        targets = ["MOLECULAR_WEIGHT", "DENSITY"]
        if slot_name in {"additive", "modifier"}:
            targets.append("SOLUTE_IDENTITY")
        res.need_traceback = True
        res.traceback_target = targets
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "molar concentration needs molecular weight and density for wt% conversion"
        return res

    # 6) v/v% 和 w/v%
    if unit_std == "v/v%":
        res.percent_type_inferred = "VOL_PERCENT"
        res.can_convert_to_wtpercent = False
        res.requires_density = True
        res.wtpercent_status = "CANNOT_CONVERT"
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "v/v% cannot be safely converted without density and formulation details"
        return res

    if unit_std == "w/v%":
        res.percent_type_inferred = "WV_PERCENT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            # w/v% = g/100mL，若 rho=1，可近似 wt%（但仍需标假设）
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "wv_percent_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "w/v% approximated as wt% under rho≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["DENSITY", "PHASE_TYPE", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "w/v% conversion requires density and solution context"
        return res

    # 其他未知单位
    res.percent_type_inferred = "NOT_PERCENT_UNIT"
    res.can_convert_to_wtpercent = False
    res.wtpercent_status = "CANNOT_CONVERT"
    res.need_traceback = True
    res.traceback_target = ["ORIGINAL_UNIT"]
    res.review_flag = 2
    res.confidence = "LOW"
    res.reason = "unit is unsupported or requires source-specific interpretation"
    return res

# =========================
# 7. DeepSeek / LLM 客户端（默认 DeepSeek）
# =========================

class LLMClient:
    def __init__(self, model: Optional[str] = None, base_url: Optional[str] = None, api_key: Optional[str] = None,
                 mock: bool = False, temperature: float = 0.0, max_retries: int = 2, sleep_seconds: float = 1.0):
        self.mock = mock
        self.temperature = temperature
        self.max_retries = max_retries
        self.sleep_seconds = sleep_seconds
        self.model = model or os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
        self.base_url = base_url or os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1")
        self.api_key = api_key or os.getenv("DEEPSEEK_API_KEY")
        self.client = None
        if not self.mock:
            if OpenAI is None:
                raise RuntimeError("未安装 openai 包（用于调用 DeepSeek 等兼容 API）。请执行 pip install openai 或使用 --mock-llm。")
            if not self.api_key:
                raise RuntimeError("未设置 DEEPSEEK_API_KEY。请设置环境变量或传入 --api-key，或使用 --mock-llm。")
            kwargs = {"api_key": self.api_key}
            if self.base_url:
                kwargs["base_url"] = self.base_url
            self.client = OpenAI(**kwargs)

    def build_user_prompt(self, row_ctx: Dict[str, Any], slot_item: Dict[str, Any], slot_cfg: Dict[str, Any]) -> str:
        current_slot = {
            "slot_name": slot_item["slot_name"],
            "slot_item_index": slot_item.get("slot_item_index", 1),
            "label_cn": slot_cfg["label_cn"],
            "phase_hint": slot_cfg["phase_hint"],
            "fixed_solute": slot_cfg["fixed_solute"],
            "raw_solute": slot_item.get("raw_solute"),
            "raw_value": slot_item.get("raw_value"),
            "raw_unit": slot_item.get("raw_unit"),
        }
        return USER_PROMPT_TEMPLATE.format(
            row_context_json=json.dumps(row_ctx, ensure_ascii=False, indent=2),
            current_slot_json=json.dumps(current_slot, ensure_ascii=False, indent=2),
            slot_specific_hint=SLOT_HINTS.get(slot_item["slot_name"], "")
        )

    def _parse_json_text(self, text: str) -> Dict[str, Any]:
        text = text.strip()
        # 去掉包裹 markdown 的意外情况
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
        return json.loads(text)

    def judge_slot(self, row_index: int, row_id: str, row_ctx: Dict[str, Any], slot_item: Dict[str, Any], slot_cfg: Dict[str, Any]) -> SlotResult:
        if self.mock:
            return mock_judge_slot(row_index, row_id, row_ctx, slot_item)

        user_prompt = self.build_user_prompt(row_ctx, slot_item, slot_cfg)
        last_err = None
        for _ in range(self.max_retries + 1):
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    temperature=self.temperature,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format={"type": "json_object"},
                )
                content = resp.choices[0].message.content
                obj = self._parse_json_text(content)
                return slot_result_from_llm_json(row_index, row_id, obj, slot_item)
            except Exception as e:
                last_err = e
                time.sleep(self.sleep_seconds)
        raise RuntimeError(f"LLM call failed after retries: {last_err}")


def slot_result_from_llm_json(row_index: int, row_id: str, obj: Dict[str, Any], slot_item: Dict[str, Any]) -> SlotResult:
    # 只做轻量 schema 容错，不做复杂修补
    def _enum(v: Any, allowed: set, default: str) -> str:
        return v if isinstance(v, str) and v in allowed else default

    traceback_target = obj.get("traceback_target") or []
    if not isinstance(traceback_target, list):
        traceback_target = []
    traceback_target = [x for x in traceback_target if x in TRACEBACK_TARGET_ALLOWED]

    return SlotResult(
        row_index=row_index,
        row_id=row_id,
        slot_name=str(obj.get("slot_name") or slot_item.get("slot_name")),
        slot_item_index=int(obj.get("slot_item_index") or slot_item.get("slot_item_index", 1)),
        parse_status=_enum(obj.get("parse_status"), PARSE_STATUS, "FAILED_PARSE"),
        mapping_status=_enum(obj.get("mapping_status"), MAPPING_STATUS, slot_item.get("mapping_status", "SINGLE_ITEM")),
        raw_solute=obj.get("raw_solute"),
        raw_value=obj.get("raw_value"),
        raw_unit=obj.get("raw_unit"),
        parsed_solute=obj.get("parsed_solute"),
        parsed_solute_alias=obj.get("parsed_solute_alias"),
        parsed_value=_coerce_optional_float(obj.get("parsed_value")),
        parsed_unit=obj.get("parsed_unit"),
        phase_identified=_enum(obj.get("phase_identified"), PHASE_STATUS, "unknown"),
        phase_evidence=_enum(obj.get("phase_evidence"), PHASE_EVIDENCE, "unknown"),
        solvent_identified=obj.get("solvent_identified"),
        percent_type_inferred=_enum(obj.get("percent_type_inferred"), PERCENT_TYPE, "NOT_PERCENT_UNIT"),
        can_convert_to_wtpercent=bool(obj.get("can_convert_to_wtpercent", False)),
        wtpercent_value=_coerce_optional_float(obj.get("wtpercent_value")),
        wtpercent_status=_enum(obj.get("wtpercent_status"), WTPERCENT_STATUS, "FAILED_PARSE"),
        conversion_method=obj.get("conversion_method"),
        requires_molecular_weight=bool(obj.get("requires_molecular_weight", False)),
        requires_density=bool(obj.get("requires_density", False)),
        molecular_weight_value=_coerce_optional_float(obj.get("molecular_weight_value")),
        molecular_weight_source=obj.get("molecular_weight_source"),
        density_value=_coerce_optional_float(obj.get("density_value")),
        density_source=obj.get("density_source"),
        need_traceback=bool(obj.get("need_traceback", False)),
        traceback_target=traceback_target,
        review_flag=int(obj.get("review_flag", 2)) if str(obj.get("review_flag", 2)) in {"0", "1", "2"} else 2,
        confidence=_enum(obj.get("confidence"), CONFIDENCE, "LOW"),
        reason=str(obj.get("reason") or ""),
    )


def _coerce_optional_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None

# =========================
# 8. 可选：回溯接口占位
# =========================

def get_traceback_text_stub(row: pd.Series, slot_result: SlotResult) -> Optional[str]:
    """
    这里先留空。你后面可以接：
    1) 原始抽取句子
    2) pdf/table 对应段落
    3) zip 解压后的原文缓存
    """
    return None

# =========================
# 9. 主流程：处理整个表
# =========================

def process_dataframe(df: pd.DataFrame, llm: LLMClient, limit: Optional[int] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if limit is not None:
        df = df.head(limit).copy()
    else:
        df = df.copy()

    all_slot_results: List[Dict[str, Any]] = []
    main_rows: List[Dict[str, Any]] = []

    # 进度条：优先 tqdm；没有则用简易输出
    try:
        from tqdm import tqdm  # type: ignore
        row_iter = tqdm(df.iterrows(), total=len(df), desc="Processing rows", unit="row")
    except Exception:
        row_iter = df.iterrows()

    for idx, row in row_iter:
        row_id = build_row_id(row, idx)
        row_ctx = row_context_dict(row, row_id)
        slot_results_for_row: List[SlotResult] = []

        for slot_cfg in SLOT_CONFIG:
            slot_items = explode_slot_items(row, slot_cfg, idx, row_id)
            for slot_item in slot_items:
                result = llm.judge_slot(idx, row_id, row_ctx, slot_item, slot_cfg)

                # 如果后面要接“回溯二次判定”，就在这里插入
                # traceback_text = get_traceback_text_stub(row, result)
                # if result.need_traceback and traceback_text:
                #     result = llm.traceback_revise(...)

                slot_results_for_row.append(result)
                all_slot_results.append(asdict(result))

        main_row = build_main_row(row, slot_results_for_row)
        main_rows.append(main_row)

    cleaned_main = pd.concat([df.reset_index(drop=True), pd.DataFrame(main_rows)], axis=1)
    slot_results_long = pd.DataFrame(all_slot_results)
    review_queue = build_review_queue(slot_results_long)
    summary = build_summary(slot_results_long)
    return cleaned_main, slot_results_long, review_queue, summary

# =========================
# 10. 主表聚合逻辑（宽表）
# =========================

def group_results_by_slot(results: List[SlotResult]) -> Dict[str, List[SlotResult]]:
    out: Dict[str, List[SlotResult]] = {}
    for r in results:
        out.setdefault(r.slot_name, []).append(r)
    for k in out:
        out[k] = sorted(out[k], key=lambda x: x.slot_item_index)
    return out


def join_values(values: Iterable[Any]) -> Optional[str]:
    vals = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        vals.append(s)
    return "; ".join(vals) if vals else None


def build_main_row(orig_row: pd.Series, results: List[SlotResult]) -> Dict[str, Any]:
    grouped = group_results_by_slot(results)
    out: Dict[str, Any] = {}

    # 每个 slot 统一输出 8 个关键字段 + 计数
    for slot_cfg in SLOT_CONFIG:
        slot_name = slot_cfg["slot_name"]
        label_cn = slot_cfg["label_cn"]
        rs = grouped.get(slot_name, [])

        if not rs:
            out[f"{label_cn}_std"] = None
            out[f"{label_cn}浓度_num"] = None
            out[f"{label_cn}浓度_unit_std"] = None
            out[f"{label_cn}_wt%"] = None
            out[f"{label_cn}_wt%_status"] = None
            out[f"{label_cn}_phase"] = None
            out[f"{label_cn}_review_flag"] = None
            out[f"{label_cn}_reason"] = None
            out[f"{label_cn}_review_count"] = 0
            continue

        out[f"{label_cn}_std"] = join_values(r.parsed_solute for r in rs)
        out[f"{label_cn}浓度_num"] = join_values(format_optional_num(r.parsed_value) for r in rs)
        out[f"{label_cn}浓度_unit_std"] = join_values(r.parsed_unit for r in rs)
        out[f"{label_cn}_wt%"] = join_values(format_optional_num(r.wtpercent_value) for r in rs)
        out[f"{label_cn}_wt%_status"] = join_values(r.wtpercent_status for r in rs)
        out[f"{label_cn}_phase"] = join_values(r.phase_identified for r in rs)
        out[f"{label_cn}_review_flag"] = max(r.review_flag for r in rs)
        out[f"{label_cn}_reason"] = join_values(r.reason for r in rs)
        out[f"{label_cn}_review_count"] = sum(1 for r in rs if r.review_flag > 0)

    out["总review_count"] = sum(v for k, v in out.items() if k.endswith("_review_count") and isinstance(v, int))
    out["建议复核字段"] = join_values(
        slot_cfg["label_cn"]
        for slot_cfg in SLOT_CONFIG
        if out.get(f"{slot_cfg['label_cn']}_review_count", 0) > 0
    )
    return out


def format_optional_num(v: Optional[float]) -> Optional[str]:
    if v is None:
        return None
    if abs(v - int(v)) < 1e-12:
        return str(int(v))
    return f"{v:.6g}"

# =========================
# 11. 异常表与统计表
# =========================

def build_review_queue(slot_results_long: pd.DataFrame) -> pd.DataFrame:
    if slot_results_long.empty:
        return slot_results_long.copy()
    mask = (
        (slot_results_long["review_flag"] > 0)
        | (~slot_results_long["wtpercent_status"].isin(["DIRECT_WT", "SAFE_CONVERTED"]))
    )
    cols = [
        "row_index", "row_id", "slot_name", "slot_item_index",
        "raw_solute", "raw_value", "raw_unit",
        "parsed_solute", "parsed_value", "parsed_unit",
        "phase_identified", "percent_type_inferred",
        "wtpercent_value", "wtpercent_status",
        "requires_molecular_weight", "requires_density",
        "need_traceback", "traceback_target",
        "review_flag", "confidence", "reason"
    ]
    cols = [c for c in cols if c in slot_results_long.columns]
    q = slot_results_long.loc[mask, cols].copy()
    q = q.sort_values(by=["review_flag", "row_index", "slot_name", "slot_item_index"], ascending=[False, True, True, True])
    return q.reset_index(drop=True)


def build_summary(slot_results_long: pd.DataFrame) -> pd.DataFrame:
    if slot_results_long.empty:
        return pd.DataFrame(columns=["metric", "value"])

    metrics = []
    metrics.append(("total_slot_items", int(len(slot_results_long))))
    for s in sorted(slot_results_long["slot_name"].dropna().unique().tolist()):
        metrics.append((f"slot_items__{s}", int((slot_results_long["slot_name"] == s).sum())))
    for s in sorted(WTPERCENT_STATUS):
        metrics.append((f"wtpercent_status__{s}", int((slot_results_long["wtpercent_status"] == s).sum())))
    for s in sorted(PARSE_STATUS):
        metrics.append((f"parse_status__{s}", int((slot_results_long["parse_status"] == s).sum())))
    metrics.append(("review_flag__1_or_2", int((slot_results_long["review_flag"] > 0).sum())))
    metrics.append(("review_flag__2", int((slot_results_long["review_flag"] == 2).sum())))
    return pd.DataFrame(metrics, columns=["metric", "value"])

# =========================
# 12. Excel 输出
# =========================

def write_excel(output_path: Path, cleaned_main: pd.DataFrame, slot_results_long: pd.DataFrame,
                review_queue: pd.DataFrame, summary: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cleaned_main.to_excel(writer, sheet_name="cleaned_main", index=False)
        slot_results_long.to_excel(writer, sheet_name="slot_results_long", index=False)
        review_queue.to_excel(writer, sheet_name="review_queue", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 自动列宽（简单版）
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for c in col[:200]:  # 避免超大表太慢，这里只估前200行
                val = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

        # review_queue 特殊上色
        if ws.title == "review_queue" and ws.max_row > 1:
            header_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
            review_col = header_map.get("review_flag")
            if review_col:
                for r in range(2, ws.max_row + 1):
                    flag_val = ws.cell(r, review_col).value
                    fill = REVIEW_FILL if flag_val in (1, 2) else OK_FILL
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = fill

    wb.save(output_path)

# =========================
# 13. CLI
# =========================

def parse_args() -> argparse.Namespace:
    cfg = RUN_CONFIG
    p = argparse.ArgumentParser(description="膜材料数据库浓度清洗脚本（LLM + 规则）")
    p.add_argument("--input", default=None, help="输入 Excel 路径；不填则用上方 RUN_CONFIG 中的 input_file（与 py 同目录）")
    p.add_argument("--output", default=None, help="输出 Excel 路径；不填则用 RUN_CONFIG 中的 output_file")
    p.add_argument("--sheet", default=None, help="工作表名；不填则用 RUN_CONFIG 或第一个 sheet")
    p.add_argument("--limit", type=int, default=None, help="只处理前 N 行；不填则用 RUN_CONFIG 中的 limit")
    p.add_argument("--mock-llm", action="store_true", default=cfg["mock_llm"], help="使用本地 mock 判定，不调用 DeepSeek；不填则用 RUN_CONFIG 中的 mock_llm")
    p.add_argument("--model", default=None, help="LLM 模型名；默认取环境变量 DEEPSEEK_MODEL")
    p.add_argument("--base-url", default=None, help="API base URL；默认取环境变量 DEEPSEEK_BASE_URL")
    p.add_argument("--api-key", default=None, help="DeepSeek API Key（或设置环境变量 DEEPSEEK_API_KEY）")
    args = p.parse_args()
    # 未在命令行指定的，用 RUN_CONFIG
    if args.input is None:
        args.input = str(SCRIPT_DIR / cfg["input_file"])
    if args.output is None:
        args.output = str(SCRIPT_DIR / cfg["output_file"])
    if args.sheet is None:
        args.sheet = cfg["sheet_name"]
    if args.limit is None:
        args.limit = cfg["limit"]
    return args


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    xls = pd.ExcelFile(input_path)
    sheet_name = args.sheet or xls.sheet_names[0]
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    required_cols = [
        "膜类型", "样品编号", "膜材料",
        "水相单体", "水相单体浓度", "水相单体浓度_单位",
        "油相单体", "油相单体浓度", "油相单体浓度_单位",
        "添加剂", "添加剂浓度", "添加剂浓度_单位",
        "改性剂", "改性剂浓度", "改性剂浓度_单位",
        "测试NaCl浓度", "测试NaCl浓度_单位",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    llm = LLMClient(
        model=args.model,
        base_url=args.base_url,
        api_key=args.api_key,
        mock=args.mock_llm,
    )

    cleaned_main, slot_results_long, review_queue, summary = process_dataframe(df, llm, limit=args.limit)
    write_excel(output_path, cleaned_main, slot_results_long, review_queue, summary)

    print(f"Done. Output written to: {output_path}")
    print(f"Rows processed: {len(cleaned_main)}")
    print(f"Slot items processed: {len(slot_results_long)}")
    if not summary.empty:
        print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
