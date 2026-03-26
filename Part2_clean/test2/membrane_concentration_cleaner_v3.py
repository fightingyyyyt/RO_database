
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
膜材料数据库浓度清洗脚本（v3：静态交付主表 + 人工核对表 + 单位总表增强版）

特点
----
1. 输出 3~4 个 sheet：
   - delivery_main：最终交付主表（静态实值，不靠公式联动）
   - concentration_review：逐浓度项核对表（你主要在这里人工修改）
   - summary：汇总统计
   - unit_catalog（可选）：如果提供单位统计表，则输出单位归类结果

2. delivery_main 的浓度/单位/状态，来自 concentration_review 的 final_* 列。
   生成时 final_* 会先复制 suggested_*。
   你人工改完第二张表后，可以用 --sync-workbook 重新同步，重建 delivery_main。

3. 单位系统做了增强：
   - 直接质量分数：wt%、w/w%、mass%、% by mass、% m/m、weight% 等
   - 质量/体积类：w/v%、wt/v%、wt/vol%、% (w/v)、g/100mL、mg/mL 等
   - 质量浓度类：g/L、mg/L、μg/L、ng/L、kg/m3、g/kg、mg/kg、ppm、ppb 等
   - 摩尔浓度类：M、mM、mol/L、mmol/L、mol/dm3、mol/m3、kmol/m3 等
   - 体积分数类：v/v%、vol%、mL/L、μL/mL 等
   - 会把大量 OCR/格式碎片做归一化；明显不是浓度单位的项会标记为 unsupported

4. 裸 "%" 的策略改进：
   - 会让 DeepSeek（或 mock）判断物质更像 solid / liquid / unknown
   - solid + 裸%：倾向 wt%，但只标 ASSUMED_CONVERTED
   - liquid + 裸%：默认不强转 wt%，而是 NEED_TRACEBACK / CANNOT_CONVERT
   - unknown：保守打标

5. 摩尔浓度换算：
   - 需要分子量；优先本地字典
   - 若安装 requests 且联网可访问 PubChem，会自动查一次并做缓存
   - 对 aqueous/test_aqueous 且无更好密度时，允许 rho≈1 的假设换算，状态标 ASSUMED_CONVERTED

运行方式
--------
A) 正常清洗
python membrane_concentration_cleaner_v3.py \
  --input data.xlsx \
  --output clean_output_v3.xlsx \
  --sheet Sheet1 \
  --llm-scope ambiguous

B) 你改完 concentration_review 后，同步重建 delivery_main
python membrane_concentration_cleaner_v3.py \
  --sync-workbook clean_output_v3.xlsx

DeepSeek 环境变量（默认）
-----------------------
DEEPSEEK_API_KEY=...
DEEPSEEK_BASE_URL=https://api.deepseek.com/v1
DEEPSEEK_MODEL=deepseek-chat
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except Exception:
    OpenAI = None  # type: ignore

try:
    import requests
except Exception:
    requests = None  # type: ignore


# ============================================================
# 本地固定配置（与本脚本同目录；默认启用，不用每次在终端输入参数）
# ============================================================
SCRIPT_DIR = Path(__file__).resolve().parent

# 设为 True 后，直接运行 `python membrane_concentration_cleaner_v3.py` 即可
USE_LOCAL_CONFIG = True
LOCAL_CONFIG = {
    # 只需填写与 py 同目录下的文件名（也可以填绝对路径）
    "input": str(SCRIPT_DIR / "test1.xlsx"),
    "output": str(SCRIPT_DIR / "test1_output_v3.xlsx"),
    "sheet": None,
    "limit": 100,
    "mock_llm": True,
    "llm_scope": "ambiguous",  # all / ambiguous / none
    "unit_inventory": None,  # 可选：str(SCRIPT_DIR / "统计单位.xlsx")
    "sync_workbook": None,
    "model": "deepseek-chat",
    "base_url": "https://api.deepseek.com/v1",
    "api_key": None,
}

# ============================================================
# 常量
# ============================================================
SLOT_CONFIG = [
    {"slot_name": "aqueous_monomer", "label_cn": "水相单体", "solute_col": "水相单体", "value_col": "水相单体浓度", "unit_col": "水相单体浓度_单位", "phase_hint": "aqueous", "fixed_solute": None},
    {"slot_name": "organic_monomer", "label_cn": "油相单体", "solute_col": "油相单体", "value_col": "油相单体浓度", "unit_col": "油相单体浓度_单位", "phase_hint": "organic", "fixed_solute": None},
    {"slot_name": "additive", "label_cn": "添加剂", "solute_col": "添加剂", "value_col": "添加剂浓度", "unit_col": "添加剂浓度_单位", "phase_hint": "unknown", "fixed_solute": None},
    {"slot_name": "modifier", "label_cn": "改性剂", "solute_col": "改性剂", "value_col": "改性剂浓度", "unit_col": "改性剂浓度_单位", "phase_hint": "unknown", "fixed_solute": None},
    {"slot_name": "test_nacl", "label_cn": "测试NaCl", "solute_col": None, "value_col": "测试NaCl浓度", "unit_col": "测试NaCl浓度_单位", "phase_hint": "test_aqueous", "fixed_solute": "NaCl"},
]

TARGET_SLOT_NAMES = [x["slot_name"] for x in SLOT_CONFIG]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EDIT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
OK_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

SUPPORTED_LLM_SCOPE = {"all", "ambiguous", "none"}
SUPPORTED_PHYSICAL_FORM = {"solid", "liquid", "unknown"}
SUPPORTED_PERCENT_TYPE = {"WT_PERCENT", "VOL_PERCENT", "WV_PERCENT", "PERCENT_UNKNOWN", "NOT_PERCENT_UNIT"}
SUPPORTED_STATUS = {"DIRECT_WT", "SAFE_CONVERTED", "ASSUMED_CONVERTED", "CANNOT_CONVERT", "NEED_TRACEBACK", "FAILED_PARSE"}
SUPPORTED_TRACEBACK_TARGETS = {"SOLUTE_IDENTITY", "PHASE_TYPE", "PERCENT_TYPE", "SOLVENT_TYPE", "MOLECULAR_WEIGHT", "DENSITY", "MAPPING_RELATION", "ORIGINAL_UNIT"}

# 常见物质分子量（g/mol）
LOCAL_MW = {
    "nacl": 58.4428,
    "sodium chloride": 58.4428,
    "mpd": 108.14,  # m-phenylenediamine
    "m-phenylenediamine": 108.14,
    "m phenylenediamine": 108.14,
    "pip": 86.14,   # piperazine
    "piperazine": 86.14,
    "tmc": 265.48,  # trimesoyl chloride
    "trimesoyl chloride": 265.48,
    "trimesic acid chloride": 265.48,
    "pei": 43.07,   # 单体重复单元近似，不推荐自动用；这里只做兜底
    "sds": 288.38,
    "sls": 288.38,
    "sodium lauryl sulfate": 288.38,
    "sodium dodecyl sulfate": 288.38,
    "camphorsulfonic acid": 232.30,
    "csa": 232.30,
    "pva": 44.05,  # 重复单元近似
    "peg": 44.05,  # 重复单元近似
}

LIQUID_KEYWORDS = {
    "water", "ethanol", "methanol", "isopropanol", "propanol", "butanol", "glycerol", "glycol",
    "dmf", "dmso", "acetone", "hexane", "heptane", "cyclohexane", "toluene", "xylene",
    "ipa", "n-hexane", "triethylamine", "tea", "acetonitrile", "formamide", "pyridine",
    "aniline", "phenol", "oil", "solvent"
}
SOLID_HINT_KEYWORDS = {
    "nacl", "chloride", "sulfate", "oxide", "graphene", "go", "cao", "tio2", "sio2", "al2o3",
    "zno", "mpd", "pip", "tmc", "piperazine", "phenylenediamine", "acid", "powder", "nanoparticle",
    "zeolite", "salt", "polymer", "resin", "particle", "dopamine hydrochloride"
}
MEMBRANE_TYPE_MAP = {
    "ro": "RO", "reverse osmosis": "RO",
    "nf": "NF", "nanofiltration": "NF",
    "uf": "UF", "ultrafiltration": "UF",
    "mf": "MF", "microfiltration": "MF",
    "md": "MD", "membrane distillation": "MD",
    "fo": "FO", "forward osmosis": "FO",
}

UNSUPPORTED_UNIT_PATTERNS = [
    r"\bratio\b", r"repeat unit", r"monomer units", r"monomers count", r"\bmolecules?\b",
    r"mequiv/g", r"\blayers?\b", r"mg/cm2", r"\bod600\b", r"\bcount\b", r"\bcycles?\b",
    r"\bmrad\b", r"\b1/s\b", r"\bat%\b", r"eq epoxy/phenoxide", r"\bppm ?h\b", r"\bppmh\b",
    r"conductivity", r"\bus/cm\b", r"\bus ?cm", r"\bμs/cm\b", r"\bms/cm\b", r"\bmho/cm\b",
    r"\bma\b", r"\bv\b", r"ion pairs", r"\bbq", r"mgo2/l", r"cn\)-/\(cd\)2\+ ratio",
    r"\bconcentration\b", r"\bna\b", r"\bn/a\b", r"unit; u?m", r"\bmol/mol\b",
]

# ============================================================
# 数据结构
# ============================================================
@dataclass
class UnitNorm:
    raw_unit: Optional[str]
    canonical_unit: Optional[str]
    unit_family: str
    convertible_family: bool
    note: str

@dataclass
class ReviewRow:
    __row_index__: int
    row_id: str
    slot_name: str
    label_cn: str
    slot_item_index: int
    mapping_status: str
    original_solute: Optional[str]
    original_value: Optional[str]
    original_unit: Optional[str]
    canonical_unit: Optional[str]
    unit_family: str
    phase: str
    solvent_identified: Optional[str]
    physical_form_inferred: str
    physical_form_evidence: str
    percent_type_inferred: str
    percent_confidence: str
    mw_value: Optional[float]
    mw_source: Optional[str]
    density_assumption: Optional[str]
    suggested_solute: Optional[str]
    suggested_value: Optional[float]
    suggested_unit: Optional[str]
    suggested_status: str
    suggested_reason: str
    final_solute: Optional[str]
    final_value: Optional[float]
    final_unit: Optional[str]
    final_status: str
    final_reason: str
    need_review: int
    traceback_target: str


# ============================================================
# 基础工具
# ============================================================
def is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() == ""


def norm_text(v: Any) -> Optional[str]:
    if is_blank(v):
        return None
    s = str(v).strip().replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def first_nonblank(row: pd.Series, cols: List[str]) -> Optional[str]:
    for c in cols:
        if c in row.index and not is_blank(row.get(c)):
            return norm_text(row.get(c))
    return None


def build_row_id(row: pd.Series, idx: int) -> str:
    rid = first_nonblank(row, ["样品编号", "RowIndex", "DOI", "文件名称", "论文题目"])
    return rid or f"row_{idx}"


def split_multi(v: Optional[str]) -> List[str]:
    if v is None:
        return []
    parts = re.split(r"[;；]+", str(v))
    return [p.strip() for p in parts if p and p.strip()]


def align_lists(solutes: List[str], values: List[str], units: List[str], fixed_solute: Optional[str]) -> Tuple[List[Tuple[Optional[str], Optional[str], Optional[str], str]], str]:
    if fixed_solute:
        solutes = [fixed_solute] * max(1, len(values), len(units))
    if not solutes and (values or units):
        solutes = [None] * max(1, len(values), len(units))
    if not values and (solutes or units):
        values = [None] * max(1, len(solutes), len(units))
    if not units and (solutes or values):
        units = [None] * max(1, len(solutes), len(values))

    counts = [len(solutes), len(values), len(units)]
    m = max(counts) if counts else 0

    def expand(lst: List[Any]) -> List[Any]:
        if len(lst) == m:
            return lst
        if len(lst) == 1:
            return lst * m
        if len(lst) == 0:
            return [None] * m
        return lst

    mapping_status = "SINGLE_ITEM"
    nonzero = [c for c in counts if c > 0]
    if len(nonzero) > 1:
        if len(set(nonzero)) == 1:
            mapping_status = "ONE_TO_ONE" if nonzero[0] > 1 else "SINGLE_ITEM"
        else:
            if len(values) == 1 and len(solutes) > 1:
                mapping_status = "ONE_VALUE_MULTI_SOLUTE"
            elif len(units) == 1 and len(solutes) > 1:
                mapping_status = "ONE_UNIT_MULTI_SOLUTE"
            else:
                mapping_status = "MULTI_UNRESOLVED"

    solutes2 = expand(solutes)
    values2 = expand(values)
    units2 = expand(units)
    if not (len(solutes2) == len(values2) == len(units2)):
        mapping_status = "MULTI_UNRESOLVED"
        m = max(len(solutes2), len(values2), len(units2))
        solutes2 = (solutes2 + [None] * m)[:m]
        values2 = (values2 + [None] * m)[:m]
        units2 = (units2 + [None] * m)[:m]

    out = list(zip(solutes2, values2, units2, [mapping_status] * len(solutes2)))
    return out, mapping_status


def parse_number(raw: Optional[str]) -> Optional[float]:
    if raw is None:
        return None
    s = str(raw).strip()
    s = s.replace(",", "")
    # 只接受单个数字或科学计数，不接受区间/比值/文本
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s):
        try:
            return float(s)
        except Exception:
            return None
    return None


def safe_float_to_text(v: Optional[float], precision: int = 6) -> Optional[str]:
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return None
    txt = f"{v:.{precision}f}".rstrip("0").rstrip(".")
    return txt if txt else "0"


def normalize_membrane_type(v: Any) -> Any:
    if is_blank(v):
        return v
    txt = norm_text(v) or ""
    parts = split_multi(txt)
    std = []
    for p in parts:
        pl = p.lower()
        found = None
        for k, val in MEMBRANE_TYPE_MAP.items():
            if k in pl:
                found = val
                break
        if found:
            std.append(found)
    if std:
        # 去重保序
        seen = set()
        out = []
        for x in std:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return "; ".join(out)
    return txt


# ============================================================
# 单位归一化
# ============================================================
def _clean_unit_for_match(u: str) -> str:
    s = u.strip().replace("％", "%").replace("µ", "μ")
    s = s.replace("−", "-").replace("–", "-").replace("⁻", "-")
    s = s.replace("·", "·").replace("⋅", "·")
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_key(u: str) -> str:
    s = _clean_unit_for_match(u).lower()
    s = s.replace(" ", "")
    s = s.replace("μ", "u")
    s = s.replace("?", "-")
    s = s.replace("^-", "-")
    s = s.replace(".", "")
    s = s.replace("_x0008_", "")
    return s




def normalize_unit_token(raw_unit: Optional[str]) -> UnitNorm:
    if raw_unit is None or str(raw_unit).strip() == "":
        return UnitNorm(raw_unit, None, "empty", False, "empty unit")

    s0 = str(raw_unit).strip()
    s = _clean_unit_for_match(s0)
    key = _norm_key(s)

    # 1) 明显不是浓度单位
    unsupported_exact = {
        "na", "n/a", "ratio", "layers", "molecules", "count", "concentration",
        "mrad", "1/s", "at%", "od600", "mequiv/g", "mg/cm2", "sccm", "cm2",
        "ma", "v", "n", "ms", "us", "mol/mol", "ionpairs",
    }
    unsupported_substrings = [
        "repeatunit", "monomerunits", "monomerscount", "epoxy/phenoxide",
        "conductivity", "ppmh", "cycles", "bq", "cn)-/(cd)2+ratio",
        "mgo2/l", "mho/cm", "us/cm", "μs/cm", "ms/cm"
    ]
    if key in unsupported_exact or any(sub in key for sub in unsupported_substrings):
        return UnitNorm(s0, s0, "unsupported", False, "unsupported or not a concentration unit")

    # 2) 质量分数
    if (
        key in {"wt%", "wt.%", "w/w%", "w/w", "mass%", "weight%"}
        or "bymass" in key
        or "%m/m" in key
        or "m/m%" in key
        or "%(w/w)" in key
        or "%(wt)" in key
        or "massfraction" in key
        or "weightof" in key
        or "basedonweight" in key
        or "wt%inaqueoussolution" in key
    ):
        return UnitNorm(s0, "wt%", "mass_fraction", True, "explicit mass fraction")

    # 3) w/v 类
    if (
        "w/v" in key or "wt/v" in key or "wt/vol" in key or "m/v" in key
        or "%w/v" in key or "%(w/v)" in key or "%(m/v)" in key
        or "g/100ml" in key
    ):
        return UnitNorm(s0, "w/v%", "mass_per_volume_percent", True, "w/v-like percent")

    # 4) v/v 类
    if "v/v" in key or "vol%" in key or "%(v/v)" in key:
        return UnitNorm(s0, "v/v%", "volume_fraction", False, "v/v-like percent")

    # 5) 体积比
    if key == "ml/l" or key == "ul/ml" or key.endswith("ml/l") or key.endswith("ul/ml"):
        return UnitNorm(s0, "mL/L", "volume_ratio", False, "volume ratio")

    # 6) ppm / ppb
    if "ppm" in key:
        return UnitNorm(s0, "ppm", "ppm", True, "ppm-like")
    if "ppb" in key:
        return UnitNorm(s0, "ppb", "ppb", True, "ppb-like")

    # 7) 质量/质量
    if key in {"g/kg", "kg/kg"}:
        return UnitNorm(s0, "g/kg", "mass_per_mass", True, "mass per mass")
    if key == "mg/kg":
        return UnitNorm(s0, "mg/kg", "mass_per_mass", True, "mass per mass")

    # 8) 质量浓度
    if key.startswith("g/l") or key in {"gl-1", "g·l-1", "gl?1", "g/dm3", "g·dm-3"}:
        return UnitNorm(s0, "g/L", "mass_per_volume", True, "g/L-like")
    if key in {"kg/m3", "kg·m-3", "kgm-3"}:
        return UnitNorm(s0, "kg/m3", "mass_per_volume", True, "kg/m3-like")
    if key.startswith("mg/l") or key in {"mgl-1", "mg·l-1", "mgl?1", "mg/dm3", "mgl^-1", "mgkg-1"}:
        return UnitNorm(s0, "mg/L" if "kg" not in key else "mg/kg", "mass_per_volume" if "kg" not in key else "mass_per_mass", True, "mg/L-like" if "kg" not in key else "mass per mass")
    if key.startswith("ug/l") or key in {"ugl-1"}:
        return UnitNorm(s0, "μg/L", "mass_per_volume", True, "μg/L-like")
    if key.startswith("ng/l") or key in {"ngl-1"}:
        return UnitNorm(s0, "ng/L", "mass_per_volume", True, "ng/L-like")
    if key.startswith("mg/ml") or key in {"mgml-1", "mgml?1"}:
        return UnitNorm(s0, "mg/mL", "mass_per_volume", True, "mg/mL-like")
    if key.startswith("ug/ml") or key in {"ugml-1"}:
        return UnitNorm(s0, "μg/mL", "mass_per_volume", True, "μg/mL-like")
    if key.startswith("g/ml") or key == "kg/l":
        return UnitNorm(s0, "g/mL" if key.startswith("g/ml") else "kg/L", "mass_per_volume", True, "g/mL-like" if key.startswith("g/ml") else "kg/L-like")

    # 9) 摩尔浓度
    if key == "m" or re.match(r"^M(\b|\s|\()", s0):
        return UnitNorm(s0, "M", "molarity", True, "M molarity")
    if key.startswith("mmol/l") or key in {"mmoll-1", "mmoll?1"}:
        return UnitNorm(s0, "mmol/L", "molarity", True, "mmol/L-like")
    if key == "mm" or re.match(r"^mM(\b|\s|\()", s0):
        return UnitNorm(s0, "mM", "molarity", True, "mM-like")
    if key in {"mmol", "mol", "mole", "mmolin11.7mlwater"}:
        return UnitNorm(s0, s0, "amount_only", False, "amount only, not concentration")
    if key.startswith("mol/l") or key in {"moll-1", "mol·l-1", "moll-1", "mol/dm3", "moldm3", "mol·dm-3"} or "monomoles" in key:
        return UnitNorm(s0, "mol/L", "molarity", True, "mol/L-like")
    if key in {"mol/m3", "molm-3", "mol·m-3"}:
        return UnitNorm(s0, "mol/m3", "molarity", True, "mol/m3-like")
    if key in {"kmol/m3", "kmolm-3"}:
        return UnitNorm(s0, "kmol/m3", "molarity", True, "kmol/m3-like")
    if key == "molal":
        return UnitNorm(s0, "molal", "molality", False, "molality needs solvent mass")
    if key in {"mol%", "mole%", "molefractionx1e-6"} or "mol%" in key or "mole%" in key or "molarratio" in key or "mole%basedon" in key:
        return UnitNorm(s0, s0, "mole_fraction", False, "mole fraction / mol%")

    # 10) only amount
    if key in {"g", "mg", "ml", "l", "mmol"}:
        return UnitNorm(s0, s0, "amount_only", False, "amount only, no denominator")

    # 11) 裸百分号
    if "%" in s0:
        return UnitNorm(s0, "%", "percent_ambiguous", False, "unclassified percent")

    return UnitNorm(s0, s0, "unknown", False, "unknown unit")


# ============================================================
# LLM 客户端（DeepSeek 默认）
# ============================================================
SYSTEM_PROMPT = """
You are helping clean membrane-material literature data.

Decide only from the provided row context and current concentration slot:
1) what the solute most likely is,
2) whether the solute is more likely a solid, liquid, or unknown under normal lab conditions,
3) what phase the slot belongs to,
4) for a bare "%" unit, whether it more likely means WT_PERCENT, VOL_PERCENT, WV_PERCENT, or remains PERCENT_UNKNOWN.

Very important:
- Be conservative.
- Do not fabricate molecular weight or density.
- Do not convert values yourself.
- If evidence is weak, return unknown / PERCENT_UNKNOWN.
- For test_nacl, solute is NaCl unless clearly contradicted.
- Output JSON only.

Allowed enum values:
phase_identified: aqueous, organic, test_aqueous, mixed, unknown
physical_form_inferred: solid, liquid, unknown
percent_type_inferred: WT_PERCENT, VOL_PERCENT, WV_PERCENT, PERCENT_UNKNOWN, NOT_PERCENT_UNIT
confidence: HIGH, MEDIUM, LOW
traceback_target items: SOLUTE_IDENTITY, PHASE_TYPE, PERCENT_TYPE, SOLVENT_TYPE, MOLECULAR_WEIGHT, DENSITY, MAPPING_RELATION, ORIGINAL_UNIT
""".strip()

USER_PROMPT_TEMPLATE = """
Row context:
{row_context_json}

Current slot:
{current_slot_json}

Return JSON with exactly these keys:
{{
  "parsed_solute": "string or null",
  "phase_identified": "enum",
  "solvent_identified": "string or null",
  "physical_form_inferred": "solid|liquid|unknown",
  "physical_form_evidence": "string",
  "percent_type_inferred": "enum",
  "confidence": "HIGH|MEDIUM|LOW",
  "need_traceback": true,
  "traceback_target": [],
  "reason": "short string"
}}

Notes:
- If current unit is not a percent-type unit, set percent_type_inferred=NOT_PERCENT_UNIT.
- For bare %, use material physical form and row context conservatively.
- For aqueous monomer, prefer the aqueous monomer field.
- For organic monomer, prefer the organic monomer field.
- For additive and modifier, do not over-assume the phase.
- Return valid JSON only.
""".strip()


def row_context_dict(row: pd.Series, row_id: str) -> Dict[str, Any]:
    return {
        "row_id": row_id,
        "membrane_type": norm_text(row.get("膜类型")),
        "membrane_material": norm_text(row.get("膜材料")),
        "membrane_structure": norm_text(row.get("膜结构标签")),
        "aqueous_monomer": norm_text(row.get("水相单体")),
        "aqueous_monomer_value": norm_text(row.get("水相单体浓度")),
        "aqueous_monomer_unit": norm_text(row.get("水相单体浓度_单位")),
        "organic_monomer": norm_text(row.get("油相单体")),
        "organic_monomer_value": norm_text(row.get("油相单体浓度")),
        "organic_monomer_unit": norm_text(row.get("油相单体浓度_单位")),
        "additive": norm_text(row.get("添加剂")),
        "additive_value": norm_text(row.get("添加剂浓度")),
        "additive_unit": norm_text(row.get("添加剂浓度_单位")),
        "modifier": norm_text(row.get("改性剂")),
        "modifier_value": norm_text(row.get("改性剂浓度")),
        "modifier_unit": norm_text(row.get("改性剂浓度_单位")),
        "test_nacl_value": norm_text(row.get("测试NaCl浓度")),
        "test_nacl_unit": norm_text(row.get("测试NaCl浓度_单位")),
        "organic_solvent": first_nonblank(row, ["有机溶剂", "油相溶剂"]),
        "aqueous_solvent": first_nonblank(row, ["水相溶剂"]),
        "other_context": first_nonblank(row, ["备注", "测试条件", "原始抽取句子", "方法"]),
    }


def infer_physical_form_heuristic(solute: Optional[str]) -> Tuple[str, str]:
    if not solute:
        return "unknown", "no solute"
    s = solute.lower()
    if any(k in s for k in LIQUID_KEYWORDS):
        return "liquid", "keyword"
    if any(k in s for k in SOLID_HINT_KEYWORDS):
        return "solid", "keyword"
    # 常见以 -ol/-one 有时是液体，但不要过度推断
    if re.search(r"(ethanol|methanol|acetone|hexane|heptane|toluene|xylene|formamide|dmf|dmso)", s):
        return "liquid", "keyword"
    return "unknown", "heuristic"


def default_phase(slot_name: str) -> str:
    return {
        "aqueous_monomer": "aqueous",
        "organic_monomer": "organic",
        "test_nacl": "test_aqueous",
    }.get(slot_name, "unknown")


class LLMClient:
    def __init__(self, model=None, base_url=None, api_key=None, mock=False, temperature=0.0, max_retries=2, sleep_seconds=1.0):
        self.mock = mock
        self.temperature = temperature
        self.max_retries = max_retries
        self.sleep_seconds = sleep_seconds
        self.model = model or os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
        self.base_url = base_url or os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1")
        self.api_key = api_key or os.getenv("DEEPSEEK_API_KEY")
        self.client = None
        if not self.mock:
            if OpenAI is None:
                raise RuntimeError("openai package is not available. Install it or use --mock-llm.")
            if not self.api_key:
                raise RuntimeError("DEEPSEEK_API_KEY is missing. Set env var or use --mock-llm.")
            self.client = OpenAI(api_key=self.api_key, base_url=self.base_url)

    def analyze_slot(self, row_context: Dict[str, Any], slot_input: Dict[str, Any]) -> Dict[str, Any]:
        if self.mock:
            return self._mock_slot_analysis(row_context, slot_input)

        user_prompt = USER_PROMPT_TEMPLATE.format(
            row_context_json=json.dumps(row_context, ensure_ascii=False, indent=2),
            current_slot_json=json.dumps(slot_input, ensure_ascii=False, indent=2),
        )

        last_err = None
        for _ in range(self.max_retries + 1):
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    temperature=self.temperature,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format={"type": "json_object"},
                )
                content = resp.choices[0].message.content
                data = json.loads(content)
                return self._sanitize_llm_output(data, slot_input)
            except Exception as e:
                last_err = e
                time.sleep(self.sleep_seconds)
        # 失败时回退 heuristic
        fallback = self._mock_slot_analysis(row_context, slot_input)
        fallback["reason"] = f"LLM failed, fallback heuristic: {last_err}"
        return fallback

    def _sanitize_llm_output(self, d: Dict[str, Any], slot_input: Dict[str, Any]) -> Dict[str, Any]:
        out = {
            "parsed_solute": d.get("parsed_solute"),
            "phase_identified": d.get("phase_identified", default_phase(slot_input["slot_name"])),
            "solvent_identified": d.get("solvent_identified"),
            "physical_form_inferred": d.get("physical_form_inferred", "unknown"),
            "physical_form_evidence": d.get("physical_form_evidence", "llm"),
            "percent_type_inferred": d.get("percent_type_inferred", "NOT_PERCENT_UNIT"),
            "confidence": d.get("confidence", "LOW"),
            "need_traceback": bool(d.get("need_traceback", False)),
            "traceback_target": d.get("traceback_target", []),
            "reason": d.get("reason", ""),
        }
        if out["phase_identified"] not in {"aqueous", "organic", "test_aqueous", "mixed", "unknown"}:
            out["phase_identified"] = default_phase(slot_input["slot_name"])
        if out["physical_form_inferred"] not in SUPPORTED_PHYSICAL_FORM:
            out["physical_form_inferred"] = "unknown"
        if out["percent_type_inferred"] not in SUPPORTED_PERCENT_TYPE:
            out["percent_type_inferred"] = "NOT_PERCENT_UNIT"
        if out["confidence"] not in {"HIGH", "MEDIUM", "LOW"}:
            out["confidence"] = "LOW"
        cleaned_targets = []
        for t in out["traceback_target"]:
            if t in SUPPORTED_TRACEBACK_TARGETS:
                cleaned_targets.append(t)
        out["traceback_target"] = cleaned_targets
        return out

    def _mock_slot_analysis(self, row_context: Dict[str, Any], slot_input: Dict[str, Any]) -> Dict[str, Any]:
        raw_solute = norm_text(slot_input.get("raw_solute"))
        raw_unit = norm_text(slot_input.get("raw_unit"))
        canonical = normalize_unit_token(raw_unit) if raw_unit else UnitNorm(None, None, "empty", False, "")
        phase = default_phase(slot_input["slot_name"])
        physical_form, evidence = infer_physical_form_heuristic(raw_solute)

        percent_type = "NOT_PERCENT_UNIT"
        confidence = "LOW"
        need_trace = False
        targets: List[str] = []
        reason = "heuristic"

        if canonical.unit_family == "mass_fraction":
            percent_type = "WT_PERCENT"
            confidence = "HIGH"
            reason = "explicit mass fraction unit"
        elif canonical.unit_family == "mass_per_volume_percent":
            percent_type = "WV_PERCENT"
            confidence = "HIGH"
            reason = "explicit w/v-like unit"
        elif canonical.unit_family in {"volume_fraction", "volume_ratio"}:
            percent_type = "VOL_PERCENT"
            confidence = "HIGH"
            reason = "volume-based unit"
        elif canonical.unit_family == "percent_ambiguous":
            if physical_form == "solid":
                percent_type = "WT_PERCENT"
                confidence = "MEDIUM"
                reason = "bare % with solid tendency"
            elif physical_form == "liquid":
                percent_type = "VOL_PERCENT"
                confidence = "MEDIUM"
                reason = "bare % with liquid tendency"
            else:
                percent_type = "PERCENT_UNKNOWN"
                confidence = "LOW"
                need_trace = True
                targets = ["PERCENT_TYPE"]
                reason = "bare % remains ambiguous"
        else:
            percent_type = "NOT_PERCENT_UNIT"

        if slot_input["slot_name"] == "test_nacl":
            phase = "test_aqueous"
            if not raw_solute:
                raw_solute = "NaCl"
            if canonical.unit_family == "percent_ambiguous":
                percent_type = "WT_PERCENT"
                confidence = "MEDIUM"
                reason = "test NaCl % usually mass-based"

        return {
            "parsed_solute": raw_solute,
            "phase_identified": phase,
            "solvent_identified": row_context.get("aqueous_solvent") if phase in {"aqueous", "test_aqueous"} else row_context.get("organic_solvent"),
            "physical_form_inferred": physical_form,
            "physical_form_evidence": evidence,
            "percent_type_inferred": percent_type,
            "confidence": confidence,
            "need_traceback": need_trace,
            "traceback_target": targets,
            "reason": reason,
        }


# ============================================================
# 分子量查找
# ============================================================
def normalize_solute_key(solute: Optional[str]) -> Optional[str]:
    if not solute:
        return None
    s = solute.lower().strip()
    s = s.replace("α", "alpha").replace("β", "beta")
    s = re.sub(r"[,\(\)\[\]]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def lookup_mw_local(solute: Optional[str]) -> Optional[float]:
    key = normalize_solute_key(solute)
    if not key:
        return None
    return LOCAL_MW.get(key)


_MW_CACHE: Dict[str, Optional[float]] = {}

def lookup_mw_pubchem(solute: Optional[str], timeout: float = 8.0) -> Optional[float]:
    if requests is None:
        return None
    key = normalize_solute_key(solute)
    if not key:
        return None
    if key in _MW_CACHE:
        return _MW_CACHE[key]
    try:
        # 先 exact name，再 general name
        url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{}/property/MolecularWeight/JSON".format(
            requests.utils.quote(key)
        )
        r = requests.get(url, timeout=timeout)
        if r.ok:
            data = r.json()
            props = data.get("PropertyTable", {}).get("Properties", [])
            if props and "MolecularWeight" in props[0]:
                mw = float(props[0]["MolecularWeight"])
                _MW_CACHE[key] = mw
                return mw
    except Exception:
        pass
    _MW_CACHE[key] = None
    return None


def lookup_mw(solute: Optional[str], allow_remote: bool = True) -> Tuple[Optional[float], Optional[str]]:
    mw = lookup_mw_local(solute)
    if mw is not None:
        return mw, "local_dict"
    if allow_remote:
        mw = lookup_mw_pubchem(solute)
        if mw is not None:
            return mw, "pubchem"
    return None, None


# ============================================================
# 单位换算
# ============================================================
def convert_to_wtpercent(
    value: Optional[float],
    unit_norm: UnitNorm,
    phase: str,
    percent_type: str,
    physical_form: str,
    solute: Optional[str],
    allow_density_assumption: bool = True,
    allow_remote_mw: bool = True,
) -> Tuple[Optional[float], str, str, Optional[float], Optional[str], Optional[str]]:
    """
    返回：
    wt_value, status, reason, mw_value, mw_source, density_assumption
    """
    if value is None:
        return None, "FAILED_PARSE", "numeric value missing", None, None, None

    fam = unit_norm.unit_family
    can = unit_norm.canonical_unit
    density_assumption = None
    mw_value = None
    mw_source = None

    # 1) 已经是 wt%
    if fam == "mass_fraction":
        return value, "DIRECT_WT", "already mass fraction", None, None, None

    # 2) 纯质量/质量
    if fam == "mass_per_mass":
        if can == "g/kg":
            return value / 10.0, "SAFE_CONVERTED", "g/kg to wt%", None, None, None
        if can == "mg/kg":
            return value / 10000.0, "SAFE_CONVERTED", "mg/kg to wt%", None, None, None

    # 3) ppm / ppb
    if fam == "ppm":
        return value / 10000.0, "ASSUMED_CONVERTED", "ppm approximated as wt% for dilute solution", None, None, "rho≈1"
    if fam == "ppb":
        return value / 10000000.0, "ASSUMED_CONVERTED", "ppb approximated as wt% for dilute solution", None, None, "rho≈1"

    # 4) 质量/体积百分比
    if fam == "mass_per_volume_percent":
        return value, "ASSUMED_CONVERTED", "w/v approximated as wt% with rho≈1", None, None, "rho≈1"

    # 5) 体积分数
    if fam in {"volume_fraction", "volume_ratio"}:
        return None, "CANNOT_CONVERT", "volume-based unit cannot be reliably converted to wt% without density data", None, None, None

    # 6) 质量浓度
    if fam == "mass_per_volume":
        density_assumption = "rho≈1"
        if can in {"g/L", "kg/m3"}:
            return value / 10.0, "ASSUMED_CONVERTED", f"{can} approximated to wt% with rho≈1", None, None, density_assumption
        if can == "mg/L":
            return value / 10000.0, "ASSUMED_CONVERTED", "mg/L approximated to wt% with rho≈1", None, None, density_assumption
        if can == "μg/L":
            return value / 10000000.0, "ASSUMED_CONVERTED", "μg/L approximated to wt% with rho≈1", None, None, density_assumption
        if can == "ng/L":
            return value / 10000000000.0, "ASSUMED_CONVERTED", "ng/L approximated to wt% with rho≈1", None, None, density_assumption
        if can == "mg/mL":
            return value * 0.1, "ASSUMED_CONVERTED", "mg/mL approximated to wt% with rho≈1", None, None, density_assumption
        if can == "μg/mL":
            return value * 0.0001, "ASSUMED_CONVERTED", "μg/mL approximated to wt% with rho≈1", None, None, density_assumption
        if can == "g/mL":
            wt = value * 100.0
            if wt > 100.0:
                return None, "NEED_TRACEBACK", "g/mL leads to >100 wt%; density/context must be checked", None, None, density_assumption
            return wt, "ASSUMED_CONVERTED", "g/mL approximated to wt% with rho≈1", None, None, density_assumption

    # 7) 摩尔浓度
    if fam == "molarity":
        mw_value, mw_source = lookup_mw(solute, allow_remote=allow_remote_mw)
        if mw_value is None:
            return None, "NEED_TRACEBACK", "molar unit needs molecular weight", None, None, None
        density_assumption = "rho≈1"
        c_mol_per_l = None
        if can == "M" or can == "mol/L":
            c_mol_per_l = value
        elif can in {"mM", "mmol/L"}:
            c_mol_per_l = value / 1000.0
        elif can == "mol/m3":
            c_mol_per_l = value / 1000.0
        elif can == "kmol/m3":
            c_mol_per_l = value
        if c_mol_per_l is None:
            return None, "NEED_TRACEBACK", "unsupported molarity variant", mw_value, mw_source, None
        # wt% = c * MW / (10 * rho)
        wt = c_mol_per_l * mw_value / 10.0
        if wt < 0 or wt > 100:
            return None, "NEED_TRACEBACK", "molar conversion out of plausible range; check density/context", mw_value, mw_source, density_assumption
        return wt, "ASSUMED_CONVERTED", "molarity converted using MW and rho≈1", mw_value, mw_source, density_assumption

    # 8) 裸 %
    if fam == "percent_ambiguous":
        if percent_type == "WT_PERCENT":
            return value, "ASSUMED_CONVERTED", "bare % treated as wt% by context", None, None, None
        if percent_type == "WV_PERCENT":
            return value, "ASSUMED_CONVERTED", "bare % treated as w/v then approximated as wt%", None, None, "rho≈1"
        if percent_type == "VOL_PERCENT":
            return None, "CANNOT_CONVERT", "bare % likely volume-based; cannot reliably convert to wt%", None, None, None
        # 没有明确 percent_type 时，用物态做辅助
        if physical_form == "solid":
            return value, "ASSUMED_CONVERTED", "bare % treated as wt% because solute tends to be solid", None, None, None
        if physical_form == "liquid":
            return None, "NEED_TRACEBACK", "bare % with liquid solute; likely v/v or mixed definition", None, None, None
        return None, "NEED_TRACEBACK", "bare % remains ambiguous", None, None, None

    # 9) 无法转
    if fam in {"molality", "mole_fraction", "amount_only", "unsupported", "unknown"}:
        return None, "CANNOT_CONVERT", f"unit family '{fam}' cannot be safely converted to wt%", None, None, None

    return None, "CANNOT_CONVERT", "conversion rule not found", None, None, None


# ============================================================
# Review 表与 Main 表构建
# ============================================================
def should_call_llm(slot_name: str, unit_norm: UnitNorm, llm_scope: str) -> bool:
    if llm_scope == "none":
        return False
    if llm_scope == "all":
        return True
    # ambiguous 模式
    return (
        unit_norm.unit_family in {"percent_ambiguous", "molarity", "unknown"}
        or slot_name in {"additive", "modifier"}
    )


def build_slot_rows_for_row(row: pd.Series, row_idx: int, llm: LLMClient, llm_scope: str, allow_remote_mw: bool = True) -> List[ReviewRow]:
    row_id = build_row_id(row, row_idx)
    row_ctx = row_context_dict(row, row_id)
    review_rows: List[ReviewRow] = []

    for cfg in SLOT_CONFIG:
        slot_name = cfg["slot_name"]
        label_cn = cfg["label_cn"]
        solute_raw = norm_text(row.get(cfg["solute_col"])) if cfg["solute_col"] else cfg["fixed_solute"]
        value_raw = norm_text(row.get(cfg["value_col"]))
        unit_raw = norm_text(row.get(cfg["unit_col"]))

        solutes = split_multi(solute_raw) if solute_raw else []
        values = split_multi(value_raw) if value_raw else []
        units = split_multi(unit_raw) if unit_raw else []
        aligned, mapping_status = align_lists(solutes, values, units, cfg["fixed_solute"])

        # 如果三者都空，跳过
        if not solutes and not values and not units and not cfg["fixed_solute"]:
            continue

        for item_idx, (s_raw, v_raw, u_raw, mapping_status_item) in enumerate(aligned, start=1):
            unit_norm = normalize_unit_token(u_raw)
            value_num = parse_number(v_raw)
            phase = default_phase(slot_name)
            solvent = row_ctx.get("aqueous_solvent") if phase in {"aqueous", "test_aqueous"} else row_ctx.get("organic_solvent")
            physical_form, pf_evidence = infer_physical_form_heuristic(s_raw)
            percent_type = "NOT_PERCENT_UNIT"
            percent_confidence = "LOW"
            traceback_targets: List[str] = []
            llm_reason = ""

            if should_call_llm(slot_name, unit_norm, llm_scope):
                slot_input = {
                    "slot_name": slot_name,
                    "label_cn": label_cn,
                    "phase_hint": cfg["phase_hint"],
                    "fixed_solute": cfg["fixed_solute"],
                    "raw_solute": s_raw,
                    "raw_value": v_raw,
                    "raw_unit": u_raw,
                }
                llm_out = llm.analyze_slot(row_ctx, slot_input)
                s_parsed = llm_out.get("parsed_solute") or s_raw
                phase = llm_out.get("phase_identified") or phase
                solvent = llm_out.get("solvent_identified") or solvent
                physical_form = llm_out.get("physical_form_inferred") or physical_form
                pf_evidence = llm_out.get("physical_form_evidence") or pf_evidence
                percent_type = llm_out.get("percent_type_inferred", "NOT_PERCENT_UNIT")
                percent_confidence = llm_out.get("confidence", "LOW")
                traceback_targets = llm_out.get("traceback_target", []) or []
                llm_reason = llm_out.get("reason", "")
                s_for_convert = s_parsed
            else:
                s_for_convert = s_raw
                # no-LLM 兜底 percent type
                if unit_norm.unit_family == "mass_fraction":
                    percent_type = "WT_PERCENT"
                    percent_confidence = "HIGH"
                elif unit_norm.unit_family == "mass_per_volume_percent":
                    percent_type = "WV_PERCENT"
                    percent_confidence = "HIGH"
                elif unit_norm.unit_family in {"volume_fraction", "volume_ratio"}:
                    percent_type = "VOL_PERCENT"
                    percent_confidence = "HIGH"
                elif unit_norm.unit_family == "percent_ambiguous":
                    if slot_name == "test_nacl":
                        percent_type = "WT_PERCENT"
                        percent_confidence = "MEDIUM"
                    elif physical_form == "solid":
                        percent_type = "WT_PERCENT"
                        percent_confidence = "MEDIUM"
                    elif physical_form == "liquid":
                        percent_type = "VOL_PERCENT"
                        percent_confidence = "MEDIUM"
                    else:
                        percent_type = "PERCENT_UNKNOWN"
                        percent_confidence = "LOW"
                else:
                    percent_type = "NOT_PERCENT_UNIT"

            wt_value, wt_status, wt_reason, mw_value, mw_source, density_assumption = convert_to_wtpercent(
                value=value_num,
                unit_norm=unit_norm,
                phase=phase,
                percent_type=percent_type,
                physical_form=physical_form,
                solute=s_for_convert,
                allow_remote_mw=allow_remote_mw,
            )

            need_review = 0
            if wt_status in {"FAILED_PARSE", "CANNOT_CONVERT", "NEED_TRACEBACK"}:
                need_review = 2
            elif wt_status == "ASSUMED_CONVERTED":
                need_review = 1

            final_solute = s_for_convert
            final_value = wt_value
            final_unit = "wt%" if wt_value is not None else unit_norm.canonical_unit
            final_status = wt_status
            final_reason = wt_reason if wt_reason else llm_reason

            rr = ReviewRow(
                __row_index__=row_idx,
                row_id=row_id,
                slot_name=slot_name,
                label_cn=label_cn,
                slot_item_index=item_idx,
                mapping_status=mapping_status_item,
                original_solute=s_raw,
                original_value=v_raw,
                original_unit=u_raw,
                canonical_unit=unit_norm.canonical_unit,
                unit_family=unit_norm.unit_family,
                phase=phase,
                solvent_identified=solvent,
                physical_form_inferred=physical_form,
                physical_form_evidence=pf_evidence,
                percent_type_inferred=percent_type,
                percent_confidence=percent_confidence,
                mw_value=mw_value,
                mw_source=mw_source,
                density_assumption=density_assumption,
                suggested_solute=s_for_convert,
                suggested_value=wt_value,
                suggested_unit="wt%" if wt_value is not None else unit_norm.canonical_unit,
                suggested_status=wt_status,
                suggested_reason=wt_reason if wt_reason else llm_reason,
                final_solute=final_solute,
                final_value=final_value,
                final_unit=final_unit,
                final_status=final_status,
                final_reason=final_reason,
                need_review=need_review,
                traceback_target="; ".join(traceback_targets),
            )
            review_rows.append(rr)

    return review_rows


def build_review_dataframe(df: pd.DataFrame, llm: LLMClient, limit: Optional[int], llm_scope: str, allow_remote_mw: bool = True) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    work_df = df.copy()
    work_df = work_df.reset_index(drop=True)
    n = len(work_df) if limit is None else min(len(work_df), limit)

    for idx in range(n):
        row = work_df.iloc[idx]
        for rr in build_slot_rows_for_row(row, idx, llm, llm_scope=llm_scope, allow_remote_mw=allow_remote_mw):
            rows.append(asdict(rr))
    return pd.DataFrame(rows)


def aggregate_slot(review_df: pd.DataFrame, row_idx: int, slot_name: str, field: str) -> Optional[str]:
    sub = review_df[(review_df["__row_index__"] == row_idx) & (review_df["slot_name"] == slot_name)]
    if sub.empty:
        return None
    vals = []
    for v in sub[field].tolist():
        if is_blank(v):
            continue
        if isinstance(v, float):
            txt = safe_float_to_text(v)
        else:
            txt = str(v).strip()
        if txt:
            vals.append(txt)
    if not vals:
        return None
    return "; ".join(vals)


def build_delivery_main(original_df: pd.DataFrame, review_df: pd.DataFrame) -> pd.DataFrame:
    main = original_df.copy().reset_index(drop=True)
    if "__row_index__" in main.columns:
        main["__row_index__"] = list(range(len(main)))
    else:
        main.insert(0, "__row_index__", list(range(len(main))))

    # 膜类型标准化：保留原列，再新增 std 列
    if "膜类型" in main.columns and "膜类型_std" not in main.columns:
        main["膜类型_std"] = main["膜类型"].apply(normalize_membrane_type)

    for cfg in SLOT_CONFIG:
        slot = cfg["slot_name"]
        if cfg["solute_col"]:
            col = cfg["solute_col"]
            agg = [aggregate_slot(review_df, i, slot, "final_solute") for i in range(len(main))]
            main[col] = [a if a is not None else main.at[i, col] for i, a in enumerate(agg)]
        vcol = cfg["value_col"]
        ucol = cfg["unit_col"]
        agg_val = [aggregate_slot(review_df, i, slot, "final_value") for i in range(len(main))]
        agg_unit = [aggregate_slot(review_df, i, slot, "final_unit") for i in range(len(main))]
        agg_status = [aggregate_slot(review_df, i, slot, "final_status") for i in range(len(main))]
        main[vcol] = [a if a is not None else main.at[i, vcol] for i, a in enumerate(agg_val)]
        main[ucol] = [a if a is not None else main.at[i, ucol] for i, a in enumerate(agg_unit)]
        main[f"{vcol}_status"] = agg_status

    # 行级汇总
    review_count = review_df.groupby("__row_index__")["need_review"].apply(lambda s: int((s > 0).sum())).to_dict()
    main["review_count"] = main["__row_index__"].map(lambda x: review_count.get(x, 0))
    main["needs_manual_review"] = main["review_count"].map(lambda x: 1 if x > 0 else 0)
    return main


def build_summary(review_df: pd.DataFrame) -> pd.DataFrame:
    if review_df.empty:
        return pd.DataFrame({"metric": ["slot_items"], "value": [0]})
    metrics = []
    metrics.append(("slot_items", len(review_df)))
    for status, cnt in review_df["final_status"].value_counts(dropna=False).items():
        metrics.append((f"status::{status}", int(cnt)))
    for fam, cnt in review_df["unit_family"].value_counts(dropna=False).items():
        metrics.append((f"unit_family::{fam}", int(cnt)))
    metrics.append(("need_review_rows", int((review_df["need_review"] > 0).sum())))
    return pd.DataFrame(metrics, columns=["metric", "value"])


def build_unit_catalog(unit_inventory_path: Optional[str]) -> Optional[pd.DataFrame]:
    if not unit_inventory_path:
        return None
    p = Path(unit_inventory_path)
    if not p.exists():
        return None
    xls = pd.ExcelFile(p)
    if not xls.sheet_names:
        return None
    df = pd.read_excel(p, sheet_name=xls.sheet_names[0])
    raw_tokens = []
    for col in df.columns:
        for val in df[col].dropna().astype(str):
            for token in split_multi(val):
                raw_tokens.append(token)
    uniq = sorted(set(raw_tokens))
    rows = []
    for tok in uniq:
        un = normalize_unit_token(tok)
        rows.append({
            "raw_unit": tok,
            "canonical_unit": un.canonical_unit,
            "unit_family": un.unit_family,
            "convertible_family": 1 if un.convertible_family else 0,
            "note": un.note,
        })
    return pd.DataFrame(rows)


# ============================================================
# Workbook 写出 / 样式 / 同步
# ============================================================
def autosize_worksheet(ws, max_width: int = 40):
    dims = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            dims[cell.column] = min(max(dims.get(cell.column, 0), length + 2), max_width)
    for col_idx, width in dims.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws, editable_cols: Optional[List[str]] = None):
    editable_cols = editable_cols or []
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for col_name in editable_cols:
        if col_name in header_map:
            cidx = header_map[col_name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, cidx).fill = EDIT_FILL

    # need_review 列高亮
    if "need_review" in header_map:
        cidx = header_map["need_review"]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, cidx).value
            if v == 2:
                ws.cell(r, cidx).fill = REVIEW_FILL
            elif v == 0:
                ws.cell(r, cidx).fill = OK_FILL

    autosize_worksheet(ws)


def write_output_workbook(output_path: Path, delivery_main: pd.DataFrame, review_df: pd.DataFrame, summary_df: pd.DataFrame, unit_catalog_df: Optional[pd.DataFrame] = None):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        delivery_main.to_excel(writer, sheet_name="delivery_main", index=False)
        review_df.to_excel(writer, sheet_name="concentration_review", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        if unit_catalog_df is not None:
            unit_catalog_df.to_excel(writer, sheet_name="unit_catalog", index=False)

    wb = load_workbook(output_path)
    style_sheet(wb["delivery_main"])
    style_sheet(
        wb["concentration_review"],
        editable_cols=["final_solute", "final_value", "final_unit", "final_status", "final_reason"]
    )
    style_sheet(wb["summary"])
    if "unit_catalog" in wb.sheetnames:
        style_sheet(wb["unit_catalog"])
    # 隐藏内部 row_index 列
    for ws_name in ["delivery_main", "concentration_review"]:
        ws = wb[ws_name]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        if "__row_index__" in header_map:
            ws.column_dimensions[get_column_letter(header_map["__row_index__"])].hidden = True
    wb.save(output_path)


def sync_workbook_from_review(workbook_path: Path, output_path: Optional[Path] = None):
    xls = pd.ExcelFile(workbook_path)
    if "delivery_main" not in xls.sheet_names or "concentration_review" not in xls.sheet_names:
        raise ValueError("Workbook must contain delivery_main and concentration_review sheets.")
    main = pd.read_excel(workbook_path, sheet_name="delivery_main")
    review = pd.read_excel(workbook_path, sheet_name="concentration_review")
    # rebuild main targeted columns from edited final_* fields
    rebuilt = build_delivery_main(main.drop(columns=[c for c in main.columns if c.endswith("_status") or c in {"review_count", "needs_manual_review"}], errors="ignore"), review)
    summary = build_summary(review)
    # preserve optional unit_catalog
    unit_catalog = None
    if "unit_catalog" in xls.sheet_names:
        unit_catalog = pd.read_excel(workbook_path, sheet_name="unit_catalog")
    out = output_path or workbook_path
    write_output_workbook(out, rebuilt, review, summary, unit_catalog)
    return out


# ============================================================
# 参数与主流程
# ============================================================
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--input", default=None, help="输入 Excel 路径")
    p.add_argument("--output", default=None, help="输出 Excel 路径")
    p.add_argument("--sheet", default=None, help="输入 sheet 名；默认第一个")
    p.add_argument("--limit", type=int, default=None, help="只试跑前 N 行")
    p.add_argument("--mock-llm", action="store_true", help="使用 mock LLM")
    p.add_argument("--llm-scope", default="ambiguous", choices=sorted(SUPPORTED_LLM_SCOPE), help="LLM 调用范围：all / ambiguous / none")
    p.add_argument("--unit-inventory", default=None, help="单位统计表路径（可选）")
    p.add_argument("--sync-workbook", default=None, help="已人工修改 concentration_review 后，重建 delivery_main")
    p.add_argument("--model", default=None, help="DeepSeek/OpenAI-compatible 模型名")
    p.add_argument("--base-url", default=None, help="DeepSeek/OpenAI-compatible base URL")
    p.add_argument("--api-key", default=None, help="API key；默认环境变量 DEEPSEEK_API_KEY")
    p.add_argument("--no-remote-mw", action="store_true", help="禁用 PubChem 在线分子量查找")
    return p.parse_args()


def get_runtime_config():
    if USE_LOCAL_CONFIG:
        class Obj: ...
        o = Obj()
        for k, v in LOCAL_CONFIG.items():
            setattr(o, k.replace("-", "_"), v)
        return o
    return parse_args()


def main():
    args = get_runtime_config()

    if getattr(args, "sync_workbook", None):
        out = sync_workbook_from_review(Path(args.sync_workbook), Path(args.output) if getattr(args, "output", None) else None)
        print(f"Synced workbook written to: {out}")
        return

    input_path = Path(args.input)
    output_path = Path(args.output)
    sheet_name = getattr(args, "sheet", None)
    limit = getattr(args, "limit", None)
    mock_llm = bool(getattr(args, "mock_llm", False))
    llm_scope = getattr(args, "llm_scope", "ambiguous")
    unit_inventory = getattr(args, "unit_inventory", None)
    allow_remote_mw = not bool(getattr(args, "no_remote_mw", False))

    if llm_scope not in SUPPORTED_LLM_SCOPE:
        raise ValueError(f"Unsupported llm_scope: {llm_scope}")

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(input_path)
    sheet_name = sheet_name or xls.sheet_names[0]
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    # 必要列检查（部分可空，但列名需存在）
    required_cols = [
        "水相单体", "水相单体浓度", "水相单体浓度_单位",
        "油相单体", "油相单体浓度", "油相单体浓度_单位",
        "添加剂", "添加剂浓度", "添加剂浓度_单位",
        "改性剂", "改性剂浓度", "改性剂浓度_单位",
        "测试NaCl浓度", "测试NaCl浓度_单位",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    llm = LLMClient(
        model=getattr(args, "model", None),
        base_url=getattr(args, "base_url", None),
        api_key=getattr(args, "api_key", None),
        mock=mock_llm,
    )

    review_df = build_review_dataframe(df, llm, limit=limit, llm_scope=llm_scope, allow_remote_mw=allow_remote_mw)
    work_df = df.reset_index(drop=True).copy()
    if limit is not None:
        work_df = work_df.iloc[:limit].copy()
    delivery_main = build_delivery_main(work_df, review_df)
    summary_df = build_summary(review_df)
    unit_catalog_df = build_unit_catalog(unit_inventory)

    write_output_workbook(output_path, delivery_main, review_df, summary_df, unit_catalog_df)

    print(f"Done. Output written to: {output_path}")
    print(f"Rows processed: {len(work_df)}")
    print(f"Concentration items in review sheet: {len(review_df)}")
    if not summary_df.empty:
        print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
