#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
膜材料数据库浓度清洗脚本（DeepSeek + 规则 + Excel 联动交付版）

本版重点改动
------------
1) 输出改为更适合交付与人工核对的两张主 sheet：
   - delivery_main：最终交付主表（保持原表结构为主，目标浓度列/单位列由 review sheet 联动）
   - concentration_review：人工核对表（保留原物质、原浓度、原单位、建议换算值、最终值、说明）
   另附 summary 统计表。

2) delivery_main 与 concentration_review 通过 Excel 公式联动：
   - 你修改 concentration_review 里的 final_* 列
   - delivery_main 中对应的物质/浓度/单位/状态会自动更新
   - 公式依赖 Excel 365 / Microsoft 365 的 FILTER + TEXTJOIN 函数

3) 对裸 "%" 的判断增加“物质是液体还是固体”的推断：
   - physical_form_inferred = solid / liquid / unknown
   - 该信息仅作为强线索，不是绝对证据
   - solid + 语境合理时，% 可倾向判为 wt%，并标记为 ASSUMED_CONVERTED
   - liquid + 语境不明时，% 倾向判为体积分数/不确定，不直接换算为 wt%
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except Exception:
    OpenAI = None  # type: ignore

# =========================
# 0. 运行配置
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent

RUN_CONFIG = {
    "input_file": "test1.xlsx",
    "output_file": "test1_output_linked.xlsx",
    "sheet_name": None,
    "limit": None,
    "mock_llm": False,
}

# =========================
# 1. 常量、状态码、列配置
# =========================

SLOT_CONFIG = [
    {
        "slot_name": "aqueous_monomer",
        "label_cn": "水相单体",
        "solute_col": "水相单体",
        "value_col": "水相单体浓度",
        "unit_col": "水相单体浓度_单位",
        "phase_hint": "aqueous",
        "fixed_solute": None,
    },
    {
        "slot_name": "organic_monomer",
        "label_cn": "油相单体",
        "solute_col": "油相单体",
        "value_col": "油相单体浓度",
        "unit_col": "油相单体浓度_单位",
        "phase_hint": "organic",
        "fixed_solute": None,
    },
    {
        "slot_name": "additive",
        "label_cn": "添加剂",
        "solute_col": "添加剂",
        "value_col": "添加剂浓度",
        "unit_col": "添加剂浓度_单位",
        "phase_hint": "unknown",
        "fixed_solute": None,
    },
    {
        "slot_name": "modifier",
        "label_cn": "改性剂",
        "solute_col": "改性剂",
        "value_col": "改性剂浓度",
        "unit_col": "改性剂浓度_单位",
        "phase_hint": "unknown",
        "fixed_solute": None,
    },
    {
        "slot_name": "test_nacl",
        "label_cn": "测试NaCl",
        "solute_col": None,
        "value_col": "测试NaCl浓度",
        "unit_col": "测试NaCl浓度_单位",
        "phase_hint": "test_aqueous",
        "fixed_solute": "NaCl",
    },
]
SLOT_BY_NAME = {x["slot_name"]: x for x in SLOT_CONFIG}

PARSE_STATUS = {
    "OK",
    "EMPTY_ALL",
    "EMPTY_VALUE",
    "EMPTY_UNIT",
    "EMPTY_SOLUTE",
    "NON_NUMERIC_VALUE",
    "MULTI_COUNT_MISMATCH",
    "UNSUPPORTED_FORMAT",
}
MAPPING_STATUS = {
    "SINGLE_ITEM",
    "ONE_TO_ONE",
    "ONE_VALUE_MULTI_SOLUTE",
    "ONE_UNIT_MULTI_SOLUTE",
    "MULTI_UNRESOLVED",
}
PHASE_STATUS = {"aqueous", "organic", "test_aqueous", "mixed", "unknown"}
PHASE_EVIDENCE = {"column_hint", "row_context", "source_text", "chemical_knowledge", "unknown"}
PERCENT_TYPE = {"WT_PERCENT", "VOL_PERCENT", "WV_PERCENT", "MOL_PERCENT", "PERCENT_UNKNOWN", "NOT_PERCENT_UNIT"}
WTPERCENT_STATUS = {"DIRECT_WT", "SAFE_CONVERTED", "ASSUMED_CONVERTED", "CANNOT_CONVERT", "NEED_TRACEBACK", "FAILED_PARSE"}
CONFIDENCE = {"HIGH", "MEDIUM", "LOW"}
TRACEBACK_TARGET_ALLOWED = {
    "SOLUTE_IDENTITY",
    "PHASE_TYPE",
    "PERCENT_TYPE",
    "SOLVENT_TYPE",
    "MOLECULAR_WEIGHT",
    "DENSITY",
    "MAPPING_RELATION",
    "ORIGINAL_UNIT",
}
PHYSICAL_FORM_ALLOWED = {"solid", "liquid", "unknown"}
PHYSICAL_FORM_EVIDENCE_ALLOWED = {"row_context", "source_text", "chemical_knowledge", "unknown"}

DIRECT_WT_UNITS = {
    "wt%", "wt.%", "w/w%", "mass%", "mass %", "g/100g", "g per 100 g", "重量%", "质量分数%"
}
PERCENT_ONLY_UNITS = {"%", "％", "percent", "percentage"}
PPM_UNITS = {"ppm", "ppm nacl"}
GL_UNITS = {"g/l", "g/L"}
MGL_UNITS = {"mg/l", "mg/L"}
VOL_UNITS = {"vol%", "v/v%", "volume%", "volume %"}
WV_UNITS = {"w/v%", "g/100ml", "g/100mL"}

LIQUID_HINT_KEYWORDS = {
    "ethanol", "methanol", "isopropanol", "propanol", "butanol", "glycerol", "glycerin",
    "acetone", "acetonitrile", "dmf", "dmac", "dmac", "nmp", "hexane", "heptane",
    "toluene", "xylene", "isopar", "cyclohexanone", "triethylamine", "tea", "aptes",
    "teos", "silane", "iptes", "mptms", "tms", "dopamine solution"
}
SOLID_HINT_KEYWORDS = {
    "nacl", "licl", "sls", "sds", "mpd", "m-phenylenediamine", "piperazine", "pip",
    "tmc", "trimesoyl chloride", "camphorsulfonic acid", "csa", "pvp", "peg", "pei",
    "dopamine hydrochloride", "go", "graphene oxide", "nanoparticle", "nanotube",
    "oxide", "zeolite", "mof", "salt", "powder", "particle", "chloride", "sulfate"
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINKED_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
EDITABLE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
OK_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

TARGET_VALUE_COLS = [cfg["value_col"] for cfg in SLOT_CONFIG]
TARGET_UNIT_COLS = [cfg["unit_col"] for cfg in SLOT_CONFIG]
TARGET_SOLUTE_COLS = [cfg["solute_col"] for cfg in SLOT_CONFIG if cfg["solute_col"] is not None]

# =========================
# 2. Prompt 模板
# =========================

SYSTEM_PROMPT = r"""
You are an expert data-cleaning assistant for membrane-material literature databases.

Your job is to normalize one concentration slot from one experimental record.

You must determine, using only the provided row context:
1. what the solute is,
2. what phase it belongs to,
3. whether the numeric value is valid,
4. what the unit means,
5. whether the concentration can be converted to wt%,
6. whether traceback to original source text is needed,
7. whether the solute is more likely a solid or a liquid in this formulation context.

Important rules:
- Do not guess missing facts.
- Do not invent molecular weight, density, solvent, or concentration meaning unless clearly supported by the provided context.
- If the evidence is insufficient, mark need_traceback=true or wtpercent_status=CANNOT_CONVERT / NEED_TRACEBACK.
- If the original unit is only "%" and its meaning is ambiguous, do not assume wt% unless the context strongly supports it.
- You should infer whether the solute is more likely solid, liquid, or unknown. Use this only as supporting evidence, not as the sole decisive rule.
- If the raw unit is "%" only:
  - a likely solid solute may support WT_PERCENT interpretation,
  - a likely liquid solute may support VOL_PERCENT interpretation,
  - but if ambiguity remains, use PERCENT_UNKNOWN.
- For test NaCl concentration, the solute is NaCl unless explicitly contradicted.
- For aqueous monomer slot, prefer the aqueous monomer column as the solute.
- For organic monomer slot, prefer the organic monomer column as the solute.
- For additive and modifier slots, use row context to infer solute-phase relation, but do not over-infer.
- Keep parsed_value numeric only if the raw value can be reliably parsed as a number.
- If multiple solutes / values / units are present, determine whether mapping is one-to-one. If mapping is unclear, mark MULTI_UNRESOLVED.
- Output JSON only.
- Do not output markdown.
- All enum values must exactly match the allowed values.

Strict anti-hallucination rule:
If a required fact is not explicitly stated or cannot be strongly inferred from the provided row context, do not fabricate it.
Prefer null, NEED_TRACEBACK, CANNOT_CONVERT, or PERCENT_UNKNOWN over unsupported completion.

Allowed enum values:

parse_status:
OK
EMPTY_ALL
EMPTY_VALUE
EMPTY_UNIT
EMPTY_SOLUTE
NON_NUMERIC_VALUE
MULTI_COUNT_MISMATCH
UNSUPPORTED_FORMAT

mapping_status:
SINGLE_ITEM
ONE_TO_ONE
ONE_VALUE_MULTI_SOLUTE
ONE_UNIT_MULTI_SOLUTE
MULTI_UNRESOLVED

phase_identified:
aqueous
organic
test_aqueous
mixed
unknown

phase_evidence:
column_hint
row_context
source_text
chemical_knowledge
unknown

percent_type_inferred:
WT_PERCENT
VOL_PERCENT
WV_PERCENT
MOL_PERCENT
PERCENT_UNKNOWN
NOT_PERCENT_UNIT

wtpercent_status:
DIRECT_WT
SAFE_CONVERTED
ASSUMED_CONVERTED
CANNOT_CONVERT
NEED_TRACEBACK
FAILED_PARSE

confidence:
HIGH
MEDIUM
LOW

review_flag:
0
1
2

physical_form_inferred:
solid
liquid
unknown

physical_form_evidence:
row_context
source_text
chemical_knowledge
unknown

traceback_target allowed items:
SOLUTE_IDENTITY
PHASE_TYPE
PERCENT_TYPE
SOLVENT_TYPE
MOLECULAR_WEIGHT
DENSITY
MAPPING_RELATION
ORIGINAL_UNIT
""".strip()

USER_PROMPT_TEMPLATE = r"""
Normalize the following concentration slot from one membrane experiment record.

Return one JSON object only.

Row context:
{row_context_json}

Current slot:
{current_slot_json}

Required output schema:
{{
  "slot_name": "string",
  "slot_item_index": 1,
  "parse_status": "enum",
  "mapping_status": "enum",
  "raw_solute": "string or null",
  "raw_value": "string or null",
  "raw_unit": "string or null",
  "parsed_solute": "string or null",
  "parsed_solute_alias": "string or null",
  "parsed_value": "number or null",
  "parsed_unit": "string or null",
  "phase_identified": "enum",
  "phase_evidence": "enum",
  "solvent_identified": "string or null",
  "physical_form_inferred": "enum",
  "physical_form_evidence": "enum",
  "percent_type_inferred": "enum",
  "can_convert_to_wtpercent": true,
  "wtpercent_value": "number or null",
  "wtpercent_status": "enum",
  "conversion_method": "string or null",
  "requires_molecular_weight": true,
  "requires_density": true,
  "molecular_weight_value": "number or null",
  "molecular_weight_source": "string or null",
  "density_value": "number or null",
  "density_source": "string or null",
  "need_traceback": true,
  "traceback_target": [],
  "review_flag": 0,
  "confidence": "enum",
  "reason": "string"
}}

Additional decision rules:
- If parse_status is not OK, then wtpercent_status should usually be FAILED_PARSE.
- If the unit is already wt%, w/w%, mass%, or equivalent mass fraction, use DIRECT_WT.
- If conversion is impossible from current evidence, use CANNOT_CONVERT or NEED_TRACEBACK.
- Do not fill molecular_weight_value or density_value unless explicitly present in the provided context.
- If the unit is only "%" and context is insufficient, set percent_type_inferred=PERCENT_UNKNOWN.
- Use physical_form_inferred only as supporting evidence, not as the only evidence.
- review_flag=0 for DIRECT_WT or SAFE_CONVERTED with strong evidence.
- review_flag=1 for ASSUMED_CONVERTED.
- review_flag=2 for FAILED_PARSE, CANNOT_CONVERT, or NEED_TRACEBACK.
- Keep reason concise, ideally within 1 sentence.

Slot-specific hint:
{slot_specific_hint}

Return valid JSON only. No explanation outside JSON.
""".strip()

SLOT_HINTS = {
    "aqueous_monomer": (
        "For aqueous_monomer, prefer the aqueous monomer column as the solute identity. "
        "Unless strong contrary evidence exists, phase_identified should be aqueous. "
        "If raw_unit is '%' only, infer WT_PERCENT only when the context strongly indicates aqueous formulation mass fraction. "
        "A likely solid monomer supports but does not guarantee WT_PERCENT."
    ),
    "organic_monomer": (
        "For organic_monomer, prefer the organic monomer column as the solute identity. "
        "Unless strong contrary evidence exists, phase_identified should be organic. "
        "Use organic_solvent context if available. "
        "Do not assume '%' means wt% unless the formulation context supports mass fraction in organic phase. "
        "A likely liquid monomer may support VOL_PERCENT, while a likely solid monomer may support WT_PERCENT."
    ),
    "additive": (
        "For additive, determine whether the additive belongs to aqueous phase, organic phase, mixed phase, or unknown based on row context. "
        "If multiple additives / values / units exist, determine whether they can be mapped one-to-one. "
        "If mapping is unclear, use mapping_status=MULTI_UNRESOLVED and need_traceback=true. "
        "Do not assume all additives are in the same phase unless supported by context. "
        "Physical form is useful but not decisive for interpreting '%' only."
    ),
    "modifier": (
        "For modifier, be conservative. "
        "A modifier may belong to post-treatment solution, coating solution, grafting solution, or another modification system. "
        "Do not force phase_identified to aqueous or organic unless supported by context. "
        "If the context is insufficient to determine the concentration meaning, set need_traceback=true. "
        "Physical form is useful but not decisive for interpreting '%' only."
    ),
    "test_nacl": (
        "For test_nacl, the solute is NaCl unless explicitly contradicted. "
        "Phase is usually test_aqueous. "
        "Units like ppm, mg/L, g/L, mol/L, and % should be interpreted as NaCl concentration in feed/test solution. "
        "NaCl is usually a solid solute, so plain '%' may support WT_PERCENT, but remain conservative if context is weak."
    ),
}

# =========================
# 3. 数据结构
# =========================

@dataclass
class SlotResult:
    row_index: int
    row_id: str
    slot_name: str
    slot_item_index: int
    parse_status: str
    mapping_status: str
    raw_solute: Optional[str]
    raw_value: Optional[str]
    raw_unit: Optional[str]
    parsed_solute: Optional[str]
    parsed_solute_alias: Optional[str]
    parsed_value: Optional[float]
    parsed_unit: Optional[str]
    phase_identified: str
    phase_evidence: str
    solvent_identified: Optional[str]
    physical_form_inferred: str
    physical_form_evidence: str
    percent_type_inferred: str
    can_convert_to_wtpercent: bool
    wtpercent_value: Optional[float]
    wtpercent_status: str
    conversion_method: Optional[str]
    requires_molecular_weight: bool
    requires_density: bool
    molecular_weight_value: Optional[float]
    molecular_weight_source: Optional[str]
    density_value: Optional[float]
    density_source: Optional[str]
    need_traceback: bool
    traceback_target: List[str]
    review_flag: int
    confidence: str
    reason: str

# =========================
# 4. 基础工具函数
# =========================

def is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    if str(v).strip() == "":
        return True
    return False


def norm_text(v: Any) -> Optional[str]:
    if is_blank(v):
        return None
    s = str(v).strip()
    s = s.replace("\u3000", " ")
    return s


def clean_unit(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    s = u.strip()
    s = s.replace("％", "%").replace("·", "/")
    s = re.sub(r"\s+", " ", s)
    return s


def canonical_unit(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    s = clean_unit(u)
    if s is None:
        return None
    s_low = s.lower().strip()
    if s_low in {x.lower() for x in DIRECT_WT_UNITS}:
        return "wt%"
    if s_low in {x.lower() for x in PERCENT_ONLY_UNITS}:
        return "%"
    if s_low in {x.lower() for x in PPM_UNITS}:
        return "ppm"
    if s_low in {x.lower() for x in GL_UNITS}:
        return "g/L"
    if s_low in {x.lower() for x in MGL_UNITS}:
        return "mg/L"
    if s_low in {x.lower() for x in VOL_UNITS}:
        return "v/v%"
    if s_low in {x.lower() for x in WV_UNITS}:
        return "w/v%"
    if s_low in {"m", "mol/l", "mol l-1", "mol·l-1"}:
        return "mol/L"
    if s == "M":
        return "M"
    if s == "mM":
        return "mM"
    return s


def split_multi(v: Optional[str]) -> List[str]:
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = re.split(r"[;；]+", s)
    return [p.strip() for p in parts if p.strip()]


def is_supported_numeric_format(s: str) -> bool:
    s = s.strip()
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s))


def to_float_safe(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    t = s.strip()
    if not is_supported_numeric_format(t):
        return None
    try:
        return float(t)
    except Exception:
        return None


def build_row_id(row: pd.Series, idx: int) -> str:
    for col in ["样品编号", "DOI", "文件名称", "论文题目"]:
        if col in row.index and not is_blank(row.get(col)):
            return str(row.get(col))
    return f"row_{idx}"


def build_other_context(row: pd.Series) -> Optional[str]:
    parts = []
    for col in ["制备方法", "过滤模式", "pH/碱度", "基底", "文件名称", "发表年份", "DOI", "论文题目"]:
        if col in row.index and not is_blank(row.get(col)):
            parts.append(f"{col}: {row.get(col)}")
    return " | ".join(parts) if parts else None


def row_context_dict(row: pd.Series, row_id: str) -> Dict[str, Any]:
    return {
        "row_id": row_id,
        "membrane_type": norm_text(row.get("膜类型")),
        "membrane_material": norm_text(row.get("膜材料")),
        "membrane_structure": norm_text(row.get("膜结构标签")) if "膜结构标签" in row.index else None,
        "aqueous_monomer": norm_text(row.get("水相单体")),
        "aqueous_monomer_value": norm_text(row.get("水相单体浓度")),
        "aqueous_monomer_unit": norm_text(row.get("水相单体浓度_单位")),
        "organic_monomer": norm_text(row.get("油相单体")),
        "organic_monomer_value": norm_text(row.get("油相单体浓度")),
        "organic_monomer_unit": norm_text(row.get("油相单体浓度_单位")),
        "additive": norm_text(row.get("添加剂")),
        "additive_value": norm_text(row.get("添加剂浓度")),
        "additive_unit": norm_text(row.get("添加剂浓度_单位")),
        "modifier": norm_text(row.get("改性剂")),
        "modifier_value": norm_text(row.get("改性剂浓度")),
        "modifier_unit": norm_text(row.get("改性剂浓度_单位")),
        "test_nacl_value": norm_text(row.get("测试NaCl浓度")),
        "test_nacl_unit": norm_text(row.get("测试NaCl浓度_单位")),
        "organic_solvent": norm_text(row.get("有机溶剂")),
        "aqueous_solvent": norm_text(row.get("水相溶剂")) if "水相溶剂" in row.index else None,
        "other_context": build_other_context(row),
    }


def format_optional_num(v: Optional[float]) -> Optional[str]:
    if v is None:
        return None
    try:
        if isinstance(v, float) and math.isnan(v):
            return None
        fv = float(v)
    except Exception:
        return None
    if abs(fv - int(fv)) < 1e-12:
        return str(int(fv))
    return f"{fv:.6g}"


def make_empty_result(row_index: int, row_id: str, slot_name: str, item_idx: int,
                      raw_solute: Optional[str], raw_value: Optional[str], raw_unit: Optional[str]) -> SlotResult:
    return SlotResult(
        row_index=row_index,
        row_id=row_id,
        slot_name=slot_name,
        slot_item_index=item_idx,
        parse_status="EMPTY_ALL",
        mapping_status="SINGLE_ITEM",
        raw_solute=raw_solute,
        raw_value=raw_value,
        raw_unit=raw_unit,
        parsed_solute=None,
        parsed_solute_alias=None,
        parsed_value=None,
        parsed_unit=None,
        phase_identified="unknown",
        phase_evidence="unknown",
        solvent_identified=None,
        physical_form_inferred="unknown",
        physical_form_evidence="unknown",
        percent_type_inferred="NOT_PERCENT_UNIT",
        can_convert_to_wtpercent=False,
        wtpercent_value=None,
        wtpercent_status="FAILED_PARSE",
        conversion_method=None,
        requires_molecular_weight=False,
        requires_density=False,
        molecular_weight_value=None,
        molecular_weight_source=None,
        density_value=None,
        density_source=None,
        need_traceback=False,
        traceback_target=[],
        review_flag=0,
        confidence="LOW",
        reason="empty slot",
    )


def join_values(values: Iterable[Any]) -> Optional[str]:
    vals = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        vals.append(s)
    return "; ".join(vals) if vals else None

# =========================
# 5. 多值拆分
# =========================

def explode_slot_items(row: pd.Series, slot_cfg: Dict[str, Any], row_index: int, row_id: str) -> List[Dict[str, Any]]:
    slot_name = slot_cfg["slot_name"]
    fixed_solute = slot_cfg["fixed_solute"]

    raw_solute = fixed_solute if fixed_solute else norm_text(row.get(slot_cfg["solute_col"]))
    raw_value = norm_text(row.get(slot_cfg["value_col"]))
    raw_unit = norm_text(row.get(slot_cfg["unit_col"]))

    if all(is_blank(x) for x in [raw_solute, raw_value, raw_unit]):
        return [{
            "slot_name": slot_name,
            "slot_item_index": 1,
            "raw_solute": raw_solute,
            "raw_value": raw_value,
            "raw_unit": raw_unit,
            "solute_count": 0,
            "value_count": 0,
            "unit_count": 0,
            "mapping_status": "SINGLE_ITEM",
        }]

    solutes = split_multi(raw_solute) if raw_solute else []
    values = split_multi(raw_value) if raw_value else []
    units = split_multi(raw_unit) if raw_unit else []

    if fixed_solute and not solutes:
        solutes = [fixed_solute]

    s_count = len(solutes)
    v_count = len(values)
    u_count = len(units)

    if max(s_count, v_count, u_count) <= 1:
        return [{
            "slot_name": slot_name,
            "slot_item_index": 1,
            "raw_solute": solutes[0] if solutes else raw_solute,
            "raw_value": values[0] if values else raw_value,
            "raw_unit": units[0] if units else raw_unit,
            "solute_count": s_count,
            "value_count": v_count,
            "unit_count": u_count,
            "mapping_status": "SINGLE_ITEM",
        }]

    if s_count == v_count == u_count and s_count > 1:
        items = []
        for i in range(s_count):
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[i],
                "raw_unit": units[i],
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_TO_ONE",
            })
        return items

    if s_count > 1 and v_count == 1 and u_count in {1, s_count}:
        items = []
        for i in range(s_count):
            item_unit = units[i] if u_count == s_count else (units[0] if u_count == 1 else raw_unit)
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[0],
                "raw_unit": item_unit,
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_VALUE_MULTI_SOLUTE",
            })
        return items

    if s_count > 1 and u_count == 1 and v_count == s_count:
        items = []
        for i in range(s_count):
            items.append({
                "slot_name": slot_name,
                "slot_item_index": i + 1,
                "raw_solute": solutes[i],
                "raw_value": values[i],
                "raw_unit": units[0],
                "solute_count": s_count,
                "value_count": v_count,
                "unit_count": u_count,
                "mapping_status": "ONE_UNIT_MULTI_SOLUTE",
            })
        return items

    return [{
        "slot_name": slot_name,
        "slot_item_index": 1,
        "raw_solute": raw_solute,
        "raw_value": raw_value,
        "raw_unit": raw_unit,
        "solute_count": s_count,
        "value_count": v_count,
        "unit_count": u_count,
        "mapping_status": "MULTI_UNRESOLVED",
    }]

# =========================
# 6. 物态推断与百分比后处理
# =========================

def infer_physical_form_simple(solute: Optional[str], slot_name: str) -> Tuple[str, str]:
    if slot_name == "test_nacl":
        return "solid", "chemical_knowledge"
    if not solute:
        return "unknown", "unknown"

    s = solute.lower().strip()
    # 更强的液体线索
    if any(k in s for k in LIQUID_HINT_KEYWORDS):
        return "liquid", "chemical_knowledge"
    # 更强的固体线索
    if any(k in s for k in SOLID_HINT_KEYWORDS):
        return "solid", "chemical_knowledge"

    # 词缀式启发
    solid_suffix_hits = ["chloride", "sulfate", "nitrate", "oxide", "nanoparticle", "powder", "salt"]
    if any(x in s for x in solid_suffix_hits):
        return "solid", "chemical_knowledge"

    liquid_suffix_hits = ["alcohol", "amine", "silane"]
    if any(x in s for x in liquid_suffix_hits):
        return "liquid", "chemical_knowledge"

    return "unknown", "unknown"


def apply_percent_bias_from_physical_form(res: SlotResult) -> SlotResult:
    unit_std = canonical_unit(res.raw_unit)
    if unit_std != "%" or res.parse_status != "OK" or res.parsed_value is None:
        return res

    # 已经明确被判成可换算时，保留原结果
    if res.wtpercent_status in {"DIRECT_WT", "SAFE_CONVERTED", "ASSUMED_CONVERTED"} and res.can_convert_to_wtpercent:
        return res

    if res.physical_form_inferred == "solid":
        # 固体溶质下，裸 % 倾向质量分数，但只给假设级
        res.percent_type_inferred = "WT_PERCENT"
        res.can_convert_to_wtpercent = True
        res.wtpercent_value = res.parsed_value
        res.wtpercent_status = "ASSUMED_CONVERTED"
        res.conversion_method = "plain_percent_to_wtpercent_solid_bias"
        res.review_flag = max(res.review_flag, 1)
        res.confidence = "MEDIUM" if res.confidence == "LOW" else res.confidence
        res.need_traceback = False if res.slot_name == "test_nacl" else res.need_traceback
        if res.reason:
            res.reason = f"{res.reason}; plain % interpreted as wt% with solid-solute bias"
        else:
            res.reason = "plain % interpreted as wt% with solid-solute bias"
        return res

    if res.physical_form_inferred == "liquid":
        res.percent_type_inferred = "VOL_PERCENT"
        res.can_convert_to_wtpercent = False
        res.wtpercent_value = None
        res.wtpercent_status = "NEED_TRACEBACK"
        res.need_traceback = True
        targets = set(res.traceback_target)
        targets.update(["PERCENT_TYPE", "SOLVENT_TYPE"])
        res.traceback_target = sorted(targets)
        res.review_flag = 2
        res.confidence = "LOW"
        if res.reason:
            res.reason = f"{res.reason}; plain % likely volume-based because solute is liquid-like"
        else:
            res.reason = "plain % likely volume-based because solute is liquid-like"
        return res

    # unknown 维持保守
    res.percent_type_inferred = "PERCENT_UNKNOWN"
    res.can_convert_to_wtpercent = False
    res.wtpercent_value = None
    res.wtpercent_status = "NEED_TRACEBACK"
    res.need_traceback = True
    targets = set(res.traceback_target)
    targets.add("PERCENT_TYPE")
    res.traceback_target = sorted(targets)
    res.review_flag = 2
    return res

# =========================
# 7. Mock 判定
# =========================

def infer_phase_by_slot(slot_name: str, row_ctx: Dict[str, Any]) -> Tuple[str, str, Optional[str]]:
    if slot_name == "aqueous_monomer":
        return "aqueous", "column_hint", (row_ctx.get("aqueous_solvent") or "water")
    if slot_name == "organic_monomer":
        return "organic", "column_hint", row_ctx.get("organic_solvent")
    if slot_name == "test_nacl":
        return "test_aqueous", "column_hint", "water"

    other_context = (row_ctx.get("other_context") or "").lower()
    organic_solvent = (row_ctx.get("organic_solvent") or "")
    if "aqueous" in other_context or "water" in other_context or "水相" in other_context:
        return "aqueous", "row_context", (row_ctx.get("aqueous_solvent") or "water")
    if organic_solvent:
        return "organic", "row_context", organic_solvent
    return "unknown", "unknown", None


def mock_judge_slot(row_index: int, row_id: str, row_ctx: Dict[str, Any], slot_item: Dict[str, Any]) -> SlotResult:
    slot_name = slot_item["slot_name"]
    raw_solute = norm_text(slot_item.get("raw_solute"))
    raw_value = norm_text(slot_item.get("raw_value"))
    raw_unit = norm_text(slot_item.get("raw_unit"))
    mapping_status = slot_item.get("mapping_status", "SINGLE_ITEM")

    res = make_empty_result(row_index, row_id, slot_name, slot_item.get("slot_item_index", 1), raw_solute, raw_value, raw_unit)
    res.mapping_status = mapping_status

    if all(is_blank(x) for x in [raw_solute, raw_value, raw_unit]):
        res.parse_status = "EMPTY_ALL"
        res.reason = "empty slot"
        res.review_flag = 0
        return res

    if slot_name != "test_nacl" and is_blank(raw_solute):
        res.parse_status = "EMPTY_SOLUTE"
        res.reason = "solute missing"
        res.review_flag = 2
        return res
    if is_blank(raw_value):
        res.parse_status = "EMPTY_VALUE"
        res.reason = "value missing"
        res.review_flag = 1 if raw_solute else 0
        return res
    if is_blank(raw_unit):
        res.parse_status = "EMPTY_UNIT"
        res.reason = "unit missing"
        res.review_flag = 2
        return res
    if mapping_status == "MULTI_UNRESOLVED":
        res.parse_status = "MULTI_COUNT_MISMATCH"
        res.need_traceback = True
        res.traceback_target = ["MAPPING_RELATION"]
        res.review_flag = 2
        res.reason = "multi-value mapping unresolved"
        return res

    parsed_num = to_float_safe(raw_value)
    unit_std = canonical_unit(raw_unit)
    phase_identified, phase_evidence, solvent_identified = infer_phase_by_slot(slot_name, row_ctx)
    physical_form, physical_form_evidence = infer_physical_form_simple(raw_solute or ("NaCl" if slot_name == "test_nacl" else None), slot_name)

    res.raw_solute = raw_solute
    res.raw_value = raw_value
    res.raw_unit = raw_unit
    res.parsed_solute = raw_solute if raw_solute else ("NaCl" if slot_name == "test_nacl" else None)
    res.parsed_solute_alias = raw_solute
    res.parsed_value = parsed_num
    res.parsed_unit = unit_std
    res.phase_identified = phase_identified
    res.phase_evidence = phase_evidence
    res.solvent_identified = solvent_identified
    res.physical_form_inferred = physical_form
    res.physical_form_evidence = physical_form_evidence

    if parsed_num is None:
        res.parse_status = "UNSUPPORTED_FORMAT" if re.search(r"[:/]|trace|few|excess|small amount|saturated", raw_value.lower()) else "NON_NUMERIC_VALUE"
        res.wtpercent_status = "FAILED_PARSE"
        res.review_flag = 2
        res.reason = "value is not reliably numeric"
        return res

    res.parse_status = "OK"

    if unit_std == "wt%":
        res.percent_type_inferred = "WT_PERCENT"
        res.can_convert_to_wtpercent = True
        res.wtpercent_value = parsed_num
        res.wtpercent_status = "DIRECT_WT"
        res.conversion_method = "direct_standardization"
        res.review_flag = 0
        res.confidence = "HIGH"
        res.reason = "original unit already expresses mass fraction"
        return res

    if unit_std == "%":
        # 先保守生成，再交由 physical_form 后处理
        res.percent_type_inferred = "PERCENT_UNKNOWN"
        res.can_convert_to_wtpercent = False
        res.wtpercent_status = "NEED_TRACEBACK"
        res.need_traceback = True
        res.traceback_target = ["PERCENT_TYPE"]
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "plain percent is ambiguous"
        return apply_percent_bias_from_physical_form(res)

    if unit_std == "ppm":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10000.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "ppm_to_wtpercent_dilute_aqueous_approx"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "ppm converted using dilute aqueous approximation"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "ppm conversion needs phase/solution context"
        return res

    if unit_std == "g/L":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "g_per_L_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "g/L converted assuming density≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "DENSITY", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "g/L conversion requires density and phase context"
        return res

    if unit_std == "mg/L":
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num / 10000.0
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "mg_per_L_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "mg/L converted assuming density≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["PHASE_TYPE", "DENSITY", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "mg/L conversion requires density and phase context"
        return res

    if unit_std in {"M", "mM", "mol/L"}:
        res.percent_type_inferred = "NOT_PERCENT_UNIT"
        res.can_convert_to_wtpercent = False
        res.requires_molecular_weight = True
        res.requires_density = True
        res.wtpercent_status = "NEED_TRACEBACK"
        targets = ["MOLECULAR_WEIGHT", "DENSITY"]
        if slot_name in {"additive", "modifier"}:
            targets.append("SOLUTE_IDENTITY")
        res.need_traceback = True
        res.traceback_target = targets
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "molar concentration needs molecular weight and density for wt% conversion"
        return res

    if unit_std == "v/v%":
        res.percent_type_inferred = "VOL_PERCENT"
        res.can_convert_to_wtpercent = False
        res.requires_density = True
        res.wtpercent_status = "CANNOT_CONVERT"
        res.review_flag = 2
        res.confidence = "LOW"
        res.reason = "v/v% cannot be safely converted without density and formulation details"
        return res

    if unit_std == "w/v%":
        res.percent_type_inferred = "WV_PERCENT"
        if phase_identified in {"aqueous", "test_aqueous"}:
            res.can_convert_to_wtpercent = True
            res.wtpercent_value = parsed_num
            res.wtpercent_status = "ASSUMED_CONVERTED"
            res.conversion_method = "wv_percent_to_wtpercent_rho1_approx"
            res.requires_density = True
            res.density_source = "assumed_water_like_density_1.0_g_per_mL"
            res.review_flag = 1
            res.confidence = "MEDIUM"
            res.reason = "w/v% approximated as wt% under rho≈1 g/mL"
        else:
            res.can_convert_to_wtpercent = False
            res.wtpercent_status = "NEED_TRACEBACK"
            res.need_traceback = True
            res.traceback_target = ["DENSITY", "PHASE_TYPE", "SOLVENT_TYPE"]
            res.review_flag = 2
            res.confidence = "LOW"
            res.reason = "w/v% conversion requires density and solution context"
        return res

    res.percent_type_inferred = "NOT_PERCENT_UNIT"
    res.can_convert_to_wtpercent = False
    res.wtpercent_status = "CANNOT_CONVERT"
    res.need_traceback = True
    res.traceback_target = ["ORIGINAL_UNIT"]
    res.review_flag = 2
    res.confidence = "LOW"
    res.reason = "unit is unsupported or requires source-specific interpretation"
    return res

# =========================
# 8. LLM 客户端
# =========================

def _coerce_optional_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


class LLMClient:
    def __init__(self, model: Optional[str] = None, base_url: Optional[str] = None, api_key: Optional[str] = None,
                 mock: bool = False, temperature: float = 0.0, max_retries: int = 2, sleep_seconds: float = 1.0):
        self.mock = mock
        self.temperature = temperature
        self.max_retries = max_retries
        self.sleep_seconds = sleep_seconds
        self.model = model or os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
        self.base_url = base_url or os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1")
        self.api_key = api_key or os.getenv("DEEPSEEK_API_KEY")
        self.client = None
        if not self.mock:
            if OpenAI is None:
                raise RuntimeError("未安装 openai 包。请执行 pip install openai 或使用 --mock-llm。")
            if not self.api_key:
                raise RuntimeError("未设置 DEEPSEEK_API_KEY。请设置环境变量或传入 --api-key，或使用 --mock-llm。")
            kwargs = {"api_key": self.api_key}
            if self.base_url:
                kwargs["base_url"] = self.base_url
            self.client = OpenAI(**kwargs)

    def build_user_prompt(self, row_ctx: Dict[str, Any], slot_item: Dict[str, Any], slot_cfg: Dict[str, Any]) -> str:
        current_slot = {
            "slot_name": slot_item["slot_name"],
            "slot_item_index": slot_item.get("slot_item_index", 1),
            "label_cn": slot_cfg["label_cn"],
            "phase_hint": slot_cfg["phase_hint"],
            "fixed_solute": slot_cfg["fixed_solute"],
            "raw_solute": slot_item.get("raw_solute"),
            "raw_value": slot_item.get("raw_value"),
            "raw_unit": slot_item.get("raw_unit"),
        }
        return USER_PROMPT_TEMPLATE.format(
            row_context_json=json.dumps(row_ctx, ensure_ascii=False, indent=2),
            current_slot_json=json.dumps(current_slot, ensure_ascii=False, indent=2),
            slot_specific_hint=SLOT_HINTS.get(slot_item["slot_name"], "")
        )

    def _parse_json_text(self, text: str) -> Dict[str, Any]:
        text = text.strip()
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
        return json.loads(text)

    def judge_slot(self, row_index: int, row_id: str, row_ctx: Dict[str, Any], slot_item: Dict[str, Any], slot_cfg: Dict[str, Any]) -> SlotResult:
        if self.mock:
            return mock_judge_slot(row_index, row_id, row_ctx, slot_item)

        user_prompt = self.build_user_prompt(row_ctx, slot_item, slot_cfg)
        last_err = None
        for _ in range(self.max_retries + 1):
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    temperature=self.temperature,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format={"type": "json_object"},
                )
                content = resp.choices[0].message.content
                obj = self._parse_json_text(content)
                result = slot_result_from_llm_json(row_index, row_id, obj, slot_item)
                result = apply_percent_bias_from_physical_form(result)
                return result
            except Exception as e:
                last_err = e
                time.sleep(self.sleep_seconds)
        raise RuntimeError(f"LLM call failed after retries: {last_err}")


def slot_result_from_llm_json(row_index: int, row_id: str, obj: Dict[str, Any], slot_item: Dict[str, Any]) -> SlotResult:
    def _enum(v: Any, allowed: set, default: str) -> str:
        return v if isinstance(v, str) and v in allowed else default

    traceback_target = obj.get("traceback_target") or []
    if not isinstance(traceback_target, list):
        traceback_target = []
    traceback_target = [x for x in traceback_target if x in TRACEBACK_TARGET_ALLOWED]

    raw_unit = obj.get("raw_unit") if obj.get("raw_unit") is not None else slot_item.get("raw_unit")
    parsed_unit = obj.get("parsed_unit") if obj.get("parsed_unit") is not None else canonical_unit(raw_unit)
    parsed_value = _coerce_optional_float(obj.get("parsed_value"))
    if parsed_value is None:
        parsed_value = to_float_safe(norm_text(obj.get("raw_value") or slot_item.get("raw_value")))

    result = SlotResult(
        row_index=row_index,
        row_id=row_id,
        slot_name=str(obj.get("slot_name") or slot_item.get("slot_name")),
        slot_item_index=int(obj.get("slot_item_index") or slot_item.get("slot_item_index", 1)),
        parse_status=_enum(obj.get("parse_status"), PARSE_STATUS, "FAILED_PARSE"),
        mapping_status=_enum(obj.get("mapping_status"), MAPPING_STATUS, slot_item.get("mapping_status", "SINGLE_ITEM")),
        raw_solute=obj.get("raw_solute"),
        raw_value=obj.get("raw_value"),
        raw_unit=raw_unit,
        parsed_solute=obj.get("parsed_solute"),
        parsed_solute_alias=obj.get("parsed_solute_alias"),
        parsed_value=parsed_value,
        parsed_unit=parsed_unit,
        phase_identified=_enum(obj.get("phase_identified"), PHASE_STATUS, "unknown"),
        phase_evidence=_enum(obj.get("phase_evidence"), PHASE_EVIDENCE, "unknown"),
        solvent_identified=obj.get("solvent_identified"),
        physical_form_inferred=_enum(obj.get("physical_form_inferred"), PHYSICAL_FORM_ALLOWED, "unknown"),
        physical_form_evidence=_enum(obj.get("physical_form_evidence"), PHYSICAL_FORM_EVIDENCE_ALLOWED, "unknown"),
        percent_type_inferred=_enum(obj.get("percent_type_inferred"), PERCENT_TYPE, "NOT_PERCENT_UNIT"),
        can_convert_to_wtpercent=bool(obj.get("can_convert_to_wtpercent", False)),
        wtpercent_value=_coerce_optional_float(obj.get("wtpercent_value")),
        wtpercent_status=_enum(obj.get("wtpercent_status"), WTPERCENT_STATUS, "FAILED_PARSE"),
        conversion_method=obj.get("conversion_method"),
        requires_molecular_weight=bool(obj.get("requires_molecular_weight", False)),
        requires_density=bool(obj.get("requires_density", False)),
        molecular_weight_value=_coerce_optional_float(obj.get("molecular_weight_value")),
        molecular_weight_source=obj.get("molecular_weight_source"),
        density_value=_coerce_optional_float(obj.get("density_value")),
        density_source=obj.get("density_source"),
        need_traceback=bool(obj.get("need_traceback", False)),
        traceback_target=traceback_target,
        review_flag=int(obj.get("review_flag", 2)) if str(obj.get("review_flag", 2)) in {"0", "1", "2"} else 2,
        confidence=_enum(obj.get("confidence"), CONFIDENCE, "LOW"),
        reason=str(obj.get("reason") or ""),
    )

    if result.physical_form_inferred == "unknown":
        pf, pe = infer_physical_form_simple(result.parsed_solute or result.raw_solute, result.slot_name)
        result.physical_form_inferred = pf
        result.physical_form_evidence = pe

    return result

# =========================
# 9. 膜类型标准化
# =========================

def standardize_membrane_type(v: Any) -> Optional[str]:
    s = norm_text(v)
    if not s:
        return None
    parts = split_multi(s) or [s]
    hits = []
    for p in parts:
        pl = p.lower()
        if "reverse osmosis" in pl or pl == "ro" or " ro " in f" {pl} ":
            hits.append("RO")
        elif "nanofiltration" in pl or pl == "nf" or " nf " in f" {pl} ":
            hits.append("NF")
        elif "ultrafiltration" in pl or pl == "uf" or " uf " in f" {pl} ":
            hits.append("UF")
        elif "microfiltration" in pl or pl == "mf" or " mf " in f" {pl} ":
            hits.append("MF")
        elif "membrane distillation" in pl or pl == "md" or " md " in f" {pl} ":
            hits.append("MD")
        elif "forward osmosis" in pl or pl == "fo" or " fo " in f" {pl} ":
            hits.append("FO")
    uniq = []
    for x in hits:
        if x not in uniq:
            uniq.append(x)
    return "; ".join(uniq) if uniq else s

# =========================
# 10. 主流程处理
# =========================

def process_dataframe(df: pd.DataFrame, llm: LLMClient, limit: Optional[int] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if limit is not None:
        df = df.head(limit).copy()
    else:
        df = df.copy()

    all_slot_results: List[Dict[str, Any]] = []

    try:
        from tqdm import tqdm  # type: ignore
        row_iter = tqdm(df.iterrows(), total=len(df), desc="Processing rows", unit="row")
    except Exception:
        row_iter = df.iterrows()

    for idx, row in row_iter:
        row_id = build_row_id(row, idx)
        row_ctx = row_context_dict(row, row_id)

        for slot_cfg in SLOT_CONFIG:
            slot_items = explode_slot_items(row, slot_cfg, idx, row_id)
            for slot_item in slot_items:
                result = llm.judge_slot(idx, row_id, row_ctx, slot_item, slot_cfg)
                all_slot_results.append(asdict(result))

    slot_results_long = pd.DataFrame(all_slot_results)
    concentration_review = build_concentration_review(slot_results_long)
    delivery_main = build_delivery_main(df, concentration_review)
    summary = build_summary(slot_results_long, concentration_review)
    return delivery_main, concentration_review, summary

# =========================
# 11. review 表与 main 表
# =========================

def build_concentration_review(slot_results_long: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    if slot_results_long.empty:
        return pd.DataFrame()

    for rec in slot_results_long.to_dict(orient="records"):
        slot_name = rec["slot_name"]
        cfg = SLOT_BY_NAME[slot_name]
        label = cfg["label_cn"]

        suggested_final_solute = rec.get("parsed_solute") or rec.get("raw_solute")
        # 建议结果：有可换算 wt% 就给 wt%；否则保留原始数值/单位，方便人工修
        if rec.get("wtpercent_value") is not None and rec.get("wtpercent_status") in {"DIRECT_WT", "SAFE_CONVERTED", "ASSUMED_CONVERTED"}:
            suggested_final_value = format_optional_num(rec.get("wtpercent_value"))
            suggested_final_unit = "wt%"
        else:
            suggested_final_value = format_optional_num(rec.get("parsed_value")) or rec.get("raw_value")
            suggested_final_unit = rec.get("parsed_unit") or rec.get("raw_unit")

        row = {
            "__row_index": rec["row_index"],
            "row_id": rec["row_id"],
            "slot_name": slot_name,
            "slot_label_cn": label,
            "slot_item_index": rec["slot_item_index"],
            "target_solute_col": cfg["solute_col"],
            "target_value_col": cfg["value_col"],
            "target_unit_col": cfg["unit_col"],
            "raw_solute": rec.get("raw_solute"),
            "raw_value": rec.get("raw_value"),
            "raw_unit": rec.get("raw_unit"),
            "parsed_solute": rec.get("parsed_solute"),
            "parsed_value": format_optional_num(rec.get("parsed_value")),
            "parsed_unit": rec.get("parsed_unit"),
            "phase_identified": rec.get("phase_identified"),
            "phase_evidence": rec.get("phase_evidence"),
            "physical_form_inferred": rec.get("physical_form_inferred"),
            "physical_form_evidence": rec.get("physical_form_evidence"),
            "percent_type_inferred": rec.get("percent_type_inferred"),
            "suggested_wtpercent_value": format_optional_num(rec.get("wtpercent_value")),
            "suggested_status": rec.get("wtpercent_status"),
            "suggested_reason": rec.get("reason"),
            "suggested_final_solute": suggested_final_solute,
            "suggested_final_value": suggested_final_value,
            "suggested_final_unit": suggested_final_unit,
            # 下面这几列是你可以手改的
            "final_solute": suggested_final_solute,
            "final_value": suggested_final_value,
            "final_unit": suggested_final_unit,
            "final_status": rec.get("wtpercent_status"),
            "final_reason": rec.get("reason"),
            "review_flag": rec.get("review_flag"),
            "confidence": rec.get("confidence"),
            "need_traceback": rec.get("need_traceback"),
            "traceback_target": "; ".join(rec.get("traceback_target") or []),
        }
        rows.append(row)

    review_df = pd.DataFrame(rows)

    ordered_cols = [
        "__row_index", "row_id", "slot_name", "slot_label_cn", "slot_item_index",
        "target_solute_col", "target_value_col", "target_unit_col",
        "raw_solute", "raw_value", "raw_unit",
        "parsed_solute", "parsed_value", "parsed_unit",
        "phase_identified", "phase_evidence",
        "physical_form_inferred", "physical_form_evidence",
        "percent_type_inferred",
        "suggested_wtpercent_value", "suggested_status", "suggested_reason",
        "suggested_final_solute", "suggested_final_value", "suggested_final_unit",
        "final_solute", "final_value", "final_unit", "final_status", "final_reason",
        "review_flag", "confidence", "need_traceback", "traceback_target"
    ]
    ordered_cols = [c for c in ordered_cols if c in review_df.columns]
    review_df = review_df[ordered_cols].sort_values(
        by=["__row_index", "slot_name", "slot_item_index"]
    ).reset_index(drop=True)
    return review_df


def build_delivery_main(orig_df: pd.DataFrame, concentration_review: pd.DataFrame) -> pd.DataFrame:
    main_df = orig_df.copy().reset_index(drop=True)
    main_df.insert(0, "__row_index", range(len(main_df)))

    # 膜类型先静态标准化
    if "膜类型" in main_df.columns:
        main_df["膜类型"] = [standardize_membrane_type(v) for v in main_df["膜类型"]]

    if concentration_review.empty:
        # 仍然给状态列
        for cfg in SLOT_CONFIG:
            label = cfg["label_cn"]
            main_df[f"{label}_status"] = None
        main_df["总review_count"] = 0
        main_df["建议人工复核"] = 0
        return main_df

    # 这些联动列后面会写入公式，先转成 object，避免 pandas dtype 警告
    for col_name in set(TARGET_SOLUTE_COLS + TARGET_VALUE_COLS + TARGET_UNIT_COLS):
        if col_name in main_df.columns:
            main_df[col_name] = main_df[col_name].astype(object)

    review_cols = list(concentration_review.columns)
    main_cols = list(main_df.columns)

    review_col_letters = {col: get_column_letter(i + 1) for i, col in enumerate(review_cols)}
    main_col_letters = {col: get_column_letter(i + 1) for i, col in enumerate(main_cols)}
    n_review = len(concentration_review) + 1
    sheet_name = "concentration_review"

    def _sheet_ref(col_name: str) -> str:
        col_letter = review_col_letters[col_name]
        return f"'{sheet_name}'!${col_letter}$2:${col_letter}${n_review}"

    def _target_join_formula(target_col_name: str, final_col_name: str, main_row_excel: int) -> str:
        row_key_main = f"${main_col_letters['__row_index']}{main_row_excel}"
        target_name_esc = target_col_name.replace('"', '""')
        return (
            f'=IFERROR(TEXTJOIN("; ",TRUE,FILTER({_sheet_ref(final_col_name)},'
            f'({_sheet_ref("__row_index")}={row_key_main})*({_sheet_ref("target_value_col")}="{target_name_esc}"))),"")'
        )

    def _target_join_formula_solute(target_col_name: str, final_col_name: str, main_row_excel: int) -> str:
        row_key_main = f"${main_col_letters['__row_index']}{main_row_excel}"
        target_name_esc = target_col_name.replace('"', '""')
        return (
            f'=IFERROR(TEXTJOIN("; ",TRUE,FILTER({_sheet_ref(final_col_name)},'
            f'({_sheet_ref("__row_index")}={row_key_main})*({_sheet_ref("target_solute_col")}="{target_name_esc}"))),"")'
        )

    def _slot_status_formula(slot_name: str, main_row_excel: int) -> str:
        row_key_main = f"${main_col_letters['__row_index']}{main_row_excel}"
        slot_name_esc = slot_name.replace('"', '""')
        return (
            f'=IFERROR(TEXTJOIN("; ",TRUE,FILTER({_sheet_ref("final_status")},'
            f'({_sheet_ref("__row_index")}={row_key_main})*({_sheet_ref("slot_name")}="{slot_name_esc}"))),"")'
        )

    def _review_count_formula(main_row_excel: int) -> str:
        row_key_main = f"${main_col_letters['__row_index']}{main_row_excel}"
        return f'=SUMPRODUCT((--({_sheet_ref("__row_index")}={row_key_main}))*({_sheet_ref("review_flag")}>0))'

    def _need_manual_formula(main_row_excel: int) -> str:
        count_formula = _review_count_formula(main_row_excel)
        return f'=IF({count_formula}>0,1,0)'

    # 预先给需要联动的列塞公式
    for i in range(len(main_df)):
        excel_row = i + 2  # header 在第1行

        # 物质列
        for col_name in TARGET_SOLUTE_COLS:
            if col_name in main_df.columns:
                main_df.at[i, col_name] = _target_join_formula_solute(col_name, "final_solute", excel_row)

        # 数值列
        for col_name in TARGET_VALUE_COLS:
            if col_name in main_df.columns:
                main_df.at[i, col_name] = _target_join_formula(col_name, "final_value", excel_row)

        # 单位列
        for col_name in TARGET_UNIT_COLS:
            if col_name in main_df.columns:
                main_df.at[i, col_name] = _target_join_formula(col_name, "final_unit", excel_row)

    # 附加少量状态列
    for cfg in SLOT_CONFIG:
        label = cfg["label_cn"]
        main_df[f"{label}_status"] = [_slot_status_formula(cfg["slot_name"], i + 2) for i in range(len(main_df))]
    main_df["总review_count"] = [_review_count_formula(i + 2) for i in range(len(main_df))]
    main_df["建议人工复核"] = [_need_manual_formula(i + 2) for i in range(len(main_df))]

    return main_df

# =========================
# 12. 统计
# =========================

def build_summary(slot_results_long: pd.DataFrame, concentration_review: pd.DataFrame) -> pd.DataFrame:
    metrics = []
    metrics.append(("total_rows", int(concentration_review["__row_index"].nunique()) if not concentration_review.empty else 0))
    metrics.append(("total_slot_items", int(len(slot_results_long))))
    for s in sorted(WTPERCENT_STATUS):
        cnt = int((slot_results_long["wtpercent_status"] == s).sum()) if not slot_results_long.empty else 0
        metrics.append((f"wtpercent_status__{s}", cnt))
    for s in sorted(PARSE_STATUS):
        cnt = int((slot_results_long["parse_status"] == s).sum()) if not slot_results_long.empty else 0
        metrics.append((f"parse_status__{s}", cnt))
    for s in sorted(PHYSICAL_FORM_ALLOWED):
        cnt = int((slot_results_long["physical_form_inferred"] == s).sum()) if not slot_results_long.empty else 0
        metrics.append((f"physical_form__{s}", cnt))
    metrics.append(("review_rows", int((concentration_review["review_flag"] > 0).sum()) if not concentration_review.empty else 0))
    return pd.DataFrame(metrics, columns=["metric", "value"])

# =========================
# 13. Excel 输出与联动样式
# =========================

def style_sheet_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, sample_rows: int = 200):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for c in col[:sample_rows]:
            val = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def write_excel(output_path: Path, delivery_main: pd.DataFrame, concentration_review: pd.DataFrame, summary: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        delivery_main.to_excel(writer, sheet_name="delivery_main", index=False)
        concentration_review.to_excel(writer, sheet_name="concentration_review", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)

    wb = load_workbook(output_path)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws_main = wb["delivery_main"]
    ws_review = wb["concentration_review"]
    ws_summary = wb["summary"]

    for ws in [ws_main, ws_review, ws_summary]:
        ws.freeze_panes = "A2"
        style_sheet_header(ws)
        auto_width(ws)

    # 隐藏 helper 列
    main_header = {c.value: c.column for c in ws_main[1]}
    review_header = {c.value: c.column for c in ws_review[1]}
    if "__row_index" in main_header:
        ws_main.column_dimensions[get_column_letter(main_header["__row_index"])].hidden = True
    if "__row_index" in review_header:
        ws_review.column_dimensions[get_column_letter(review_header["__row_index"])].hidden = True

    # 主表里联动列上色
    linked_cols = set(TARGET_SOLUTE_COLS + TARGET_VALUE_COLS + TARGET_UNIT_COLS + [f"{cfg['label_cn']}_status" for cfg in SLOT_CONFIG])
    for col_name, idx in main_header.items():
        if col_name in linked_cols:
            col_letter = get_column_letter(idx)
            for row in range(2, ws_main.max_row + 1):
                ws_main[f"{col_letter}{row}"].fill = LINKED_FILL

    # review 表里 final_* 为可编辑区，高亮
    editable_cols = {"final_solute", "final_value", "final_unit", "final_status", "final_reason"}
    for col_name, idx in review_header.items():
        if col_name in editable_cols:
            col_letter = get_column_letter(idx)
            for row in range(2, ws_review.max_row + 1):
                ws_review[f"{col_letter}{row}"].fill = EDITABLE_FILL

    # review_flag 着色
    review_flag_col = review_header.get("review_flag")
    if review_flag_col:
        for row in range(2, ws_review.max_row + 1):
            flag_val = ws_review.cell(row, review_flag_col).value
            fill = REVIEW_FILL if flag_val in (1, 2) else OK_FILL
            for c in range(1, ws_review.max_column + 1):
                if ws_review.cell(row, c).fill == EDITABLE_FILL:
                    continue
                ws_review.cell(row, c).fill = fill

    wb.save(output_path)

# =========================
# 14. CLI
# =========================

def parse_args() -> argparse.Namespace:
    cfg = RUN_CONFIG
    p = argparse.ArgumentParser(description="膜材料数据库浓度清洗脚本（DeepSeek + Excel 联动交付版）")
    p.add_argument("--input", default=None, help="输入 Excel 路径；不填则用 RUN_CONFIG 中的 input_file")
    p.add_argument("--output", default=None, help="输出 Excel 路径；不填则用 RUN_CONFIG 中的 output_file")
    p.add_argument("--sheet", default=None, help="工作表名；不填则用 RUN_CONFIG 或第一个 sheet")
    p.add_argument("--limit", type=int, default=None, help="只处理前 N 行；不填则用 RUN_CONFIG 中的 limit")
    p.add_argument("--mock-llm", action="store_true", default=cfg["mock_llm"], help="使用本地 mock 判定，不调用 DeepSeek")
    p.add_argument("--model", default=None, help="LLM 模型名；默认取环境变量 DEEPSEEK_MODEL")
    p.add_argument("--base-url", default=None, help="API base URL；默认取环境变量 DEEPSEEK_BASE_URL")
    p.add_argument("--api-key", default=None, help="DeepSeek API Key（或设置环境变量 DEEPSEEK_API_KEY）")
    args = p.parse_args()
    if args.input is None:
        args.input = str(SCRIPT_DIR / cfg["input_file"])
    if args.output is None:
        args.output = str(SCRIPT_DIR / cfg["output_file"])
    if args.sheet is None:
        args.sheet = cfg["sheet_name"]
    if args.limit is None:
        args.limit = cfg["limit"]
    return args


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    xls = pd.ExcelFile(input_path)
    sheet_name = args.sheet or xls.sheet_names[0]
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    required_cols = [
        "膜类型", "样品编号", "膜材料",
        "水相单体", "水相单体浓度", "水相单体浓度_单位",
        "油相单体", "油相单体浓度", "油相单体浓度_单位",
        "添加剂", "添加剂浓度", "添加剂浓度_单位",
        "改性剂", "改性剂浓度", "改性剂浓度_单位",
        "测试NaCl浓度", "测试NaCl浓度_单位",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    llm = LLMClient(
        model=args.model,
        base_url=args.base_url,
        api_key=args.api_key,
        mock=args.mock_llm,
    )

    delivery_main, concentration_review, summary = process_dataframe(df, llm, limit=args.limit)
    write_excel(output_path, delivery_main, concentration_review, summary)

    print(f"Done. Output written to: {output_path}")
    print(f"Rows processed: {len(delivery_main)}")
    print(f"Review rows: {len(concentration_review)}")
    if not summary.empty:
        print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
